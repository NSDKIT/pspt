import pandas as pd
from openpyxl import load_workbook

def generate_combinations(start, end, group_size):
    # 数字の範囲を定義
    numbers = range(start, end + 1)

    # 重複組み合わせを生成
    combinations = [(i, j) for i in numbers for j in numbers if i <= j]

    # 組み合わせをグループに分ける
    groups = [combinations[i:i + group_size] for i in range(0, len(combinations), group_size)]

    # 最後のグループを補完するために、最初のグループから必要な数を追加
    if len(groups[-1]) < group_size:
        needed = group_size - len(groups[-1])
        groups[-1].extend(groups[0][:needed])

    # 各グループをDataFrameに変換
    df_list = []
    for idx, group in enumerate(groups):
        df = pd.DataFrame(group, columns=['Element1', 'Element2'])
        df_list.append(df)

    return df_list

def save_to_existing_excel(df, file_path, sheet_name, start_row, start_col):
    # 既存のエクセルファイルを読み込む
    book = load_workbook(file_path, keep_vba=True)
    sheet = book[sheet_name]

    # データを書き込む
    for i, row in df.iterrows():
        for j, value in enumerate(row):
            sheet.cell(row=start_row + i + 1, column=start_col + j + 1, value=value)

    # ファイルを保存
    book.save(file_path)

# 関数を使用して重複組み合わせを生成
df_list = generate_combinations(11, 47, 10)

# 最初のグループをエクセルに書き込むためのデータフレームを選択
df_first_group = df_list[0]

# ファイルパスと書き込み設定
file_path = 'C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\02_MyModel\\CPATFreeR05V1.0\\CPATFree_R05_Ver1.0\\model\\EAST10\\main\\condition00.xlsm'
sheet_name = 'Sheet1'  # 書き込みたいシート名
start_row = 1
start_col = 9

# 既存のエクセルファイルにデータを書き込む
save_to_existing_excel(df_first_group, excel_file_path, sheet_name, start_row, start_col)
