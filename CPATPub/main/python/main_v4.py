# =============================================================================
# CPAT 自動解析システム メインスクリプト (main_v4.py)
# =============================================================================
# 【使い方】
#   1. config.ini を開き、フォルダパスや解析条件を自分の環境に合わせて編集する
#   2. このスクリプトを実行する:  python main_v4.py
#
# 【注意】
#   実行中はマウス・キーボードに触れないでください。
#   GUI が自動操作されます。
# =============================================================================

import shutil
import os
import time
import json
import configparser
import subprocess
import zipfile
import math

from pynput import mouse, keyboard
import pygetwindow as gw
import pyautogui
import pyperclip
from openpyxl import load_workbook
import pandas as pd
import win32com.client
from itertools import combinations_with_replacement
import numpy as np

# =============================================================================
# 設定ファイルの読み込み
# =============================================================================

# このスクリプト自身が置かれているフォルダを基点とする（相対パスの基準）
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def _resolve(path: str) -> str:
    """config.ini に書かれたパスを絶対パスに変換する。
    相対パス（./ や ../ で始まる）は _BASE_DIR を基点に解決する。"""
    if not os.path.isabs(path):
        path = os.path.normpath(os.path.join(_BASE_DIR, path))
    return path

def load_config(config_path: str = None) -> configparser.ConfigParser:
    """config.ini を読み込んで返す。"""
    if config_path is None:
        config_path = os.path.join(_BASE_DIR, "config.ini")
    cfg = configparser.ConfigParser()
    cfg.read(config_path, encoding="utf-8")
    return cfg


# =============================================================================
# ファイル・フォルダ操作ユーティリティ
# =============================================================================

def rename_files(original_file: str, new_file: str):
    """ファイルの名前を変更する。"""
    os.rename(original_file, new_file)


def create_directory(directory: str):
    """ディレクトリを作成する（既に存在しても無視）。"""
    os.makedirs(directory, exist_ok=True)


def unzip_file(zip_file: str, extract_to: str):
    """ZIP ファイルを指定フォルダに解凍する。"""
    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
        zip_ref.extractall(extract_to)


def delete_files(directory: str, extension: str):
    """ディレクトリ内の指定拡張子のファイルをすべて削除する。"""
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith(extension):
                os.remove(os.path.join(root, file))


def move_files(source_directory: str, destination_directory: str, extension: str):
    """ディレクトリ内の指定拡張子のファイルを別フォルダに移動する。"""
    for root, dirs, files in os.walk(source_directory):
        for file in files:
            if file.endswith(extension):
                shutil.move(os.path.join(root, file), destination_directory)


def replace_content_in_files(directory: str, target: str, replacement: str, extension: str):
    """ディレクトリ内の指定拡張子ファイルの文字列を一括置換する。"""
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            if file.endswith(extension):
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                modified_content = content.replace(target, replacement)
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(modified_content)


def calculate_value(factor: float, res: float) -> float:
    """係数と RES 比率から値を計算する。"""
    return factor * res


def modify_demand_file(file_path: str, value_text_pgo: str, value_text_qgo: str):
    """Demand.pnsj ファイルの Pgo / Qgo タグを新しい値に書き換える。"""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    modified_content = content.replace("<Pgo>0.0</Pgo>", value_text_pgo)
    modified_content = modified_content.replace("<Qgo>0.0</Qgo>", value_text_qgo)
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(modified_content)


def replace_tags_in_file(file_path: str, old_values_pgo, new_values_pgo,
                          old_values_qgo_temp, new_values_qgo):
    """<Pgo> タグと <Qgo> タグを新しい値で置き換える。"""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    for old_value, new_value in zip(old_values_pgo, new_values_pgo):
        content = content.replace(f"<Pgo>{old_value}</Pgo>", f"<Pgo>{new_value}</Pgo>")
    for old_value, new_value in zip(old_values_qgo_temp, new_values_qgo):
        content = content.replace(f"<Qgo>{old_value}</Qgo>", f"<Qgo>{new_value}</Qgo>")
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(content)


def modify_yco_value(file_path: str, bus_value: int, yco_value: float):
    """指定バス番号の Yco 値を書き換える。"""
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    lines = content.split('\n')
    for i, line in enumerate(lines):
        if f"<Number>{bus_value}</Number>" in line:
            if i + 7 < len(lines) and "<Yco>" in lines[i + 7]:
                lines[i + 7] = f"          <Yco>{yco_value}</Yco>"
    with open(file_path, 'w', encoding='utf-8') as file:
        file.write('\n'.join(lines))
    print(f"{os.path.basename(file_path)}: バス {bus_value} の <Yco> を {yco_value} に変更しました。")


def compress_directory(directory: str, output_file: str):
    """ディレクトリを ZIP 圧縮する。"""
    shutil.make_archive(output_file.replace('.zip', ''), 'zip', directory)


def clean_up(*files_and_directories):
    """中間ファイル・作業ディレクトリを削除する。"""
    for item in files_and_directories:
        if os.path.isfile(item):
            os.remove(item)
        elif os.path.isdir(item):
            shutil.rmtree(item)


def copy_file(folder_path: str, file_name: str):
    """ファイルを同フォルダ内に _copy という名前でコピーする。"""
    source_path = os.path.join(folder_path, file_name)
    destination_path = os.path.join(folder_path, file_name.replace(".pop", "_copy.pop"))
    shutil.copy2(source_path, destination_path)
    print(f"コピー: {file_name} → {destination_path}")


def delete_file(file_path: str):
    """ファイルを削除する（存在しない場合は警告のみ）。"""
    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"削除: {file_path}")
    else:
        print(f"[警告] ファイルが見つかりません: {file_path}")


def delete_all_files_in_folder(folder_path: str):
    """フォルダ内のすべてのファイル・サブフォルダを削除する。"""
    if os.path.exists(folder_path):
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print(f'[警告] 削除失敗: {file_path}  理由: {e}')
    else:
        print(f'[警告] フォルダが存在しません: {folder_path}')


def rename_folder(old_folder_name: str, new_folder_name: str):
    """フォルダ名を変更する。"""
    try:
        os.rename(old_folder_name, new_folder_name)
        print(f"フォルダ名変更: '{old_folder_name}' → '{new_folder_name}'")
    except Exception as e:
        print(f"[エラー] フォルダ名変更失敗: {e}")


def copy_and_rename_condition_file(contingency_case: int, target_of_number: int,
                                    source_folder: str, destination_folder: str):
    """事故種別・箇所数に応じた条件テンプレートを condition0.xlsx としてコピーする。"""
    mapping = {
        (1, 1): "Condition(L)_1site.xlsx",
        (1, 2): "Condition(L)_2site.xlsx",
        (2, 1): "Condition(G_O_C)_1site.xlsx",
        (2, 2): "Condition(G_O_C)_2site.xlsx",
    }
    src_file = mapping.get((contingency_case, target_of_number))
    if src_file is None:
        raise ValueError("ContingencyCase / TargetOfNumber の組み合わせが無効です。")
    src_path = os.path.join(source_folder, src_file)
    dst_path = os.path.join(destination_folder, "condition0.xlsx")
    if os.path.exists(src_path):
        shutil.copy(src_path, dst_path)
    else:
        raise FileNotFoundError(f"条件テンプレートが見つかりません: {src_path}")


def copy_files2(main_folder: str, condition_list_folder: str,
                src_file_name: str, dst_file_name: str):
    """MDL ファイルをメインフォルダにコピーし、既存ファイルをバックアップする。"""
    src_mdl_path = os.path.join(condition_list_folder, src_file_name)
    dst_main_path = os.path.join(main_folder, dst_file_name)
    backup_mdl_path = os.path.join(condition_list_folder, dst_file_name)
    if os.path.exists(src_mdl_path):
        if os.path.exists(dst_main_path):
            shutil.move(dst_main_path, backup_mdl_path)
        shutil.copy(src_mdl_path, dst_main_path)
        os.rename(dst_main_path, os.path.join(main_folder, dst_file_name))


# =============================================================================
# プロセス管理ユーティリティ
# =============================================================================

def get_process_ids(process_name: str) -> list:
    """指定プロセス名の PID 一覧を返す。"""
    result = subprocess.run(['tasklist'], capture_output=True, text=True)
    pids = []
    for line in result.stdout.splitlines():
        if process_name.lower() in line.lower():
            parts = line.split()
            try:
                pids.append(int(parts[1]))
            except (IndexError, ValueError):
                pass
    return pids


def kill_process_v1(pid: int):
    """PID を指定してプロセスを強制終了する。"""
    subprocess.run(['taskkill', '/PID', str(pid), '/F'])


# =============================================================================
# Excel 操作ユーティリティ
# =============================================================================

def open_save_close_excel(file_path: str):
    """Excel を COM 経由で開き、VBA を実行して保存・終了する。"""
    excel_app = win32com.client.Dispatch("Excel.Application")
    excel_app.Visible = False
    excel_app.DisplayAlerts = False
    workbook = excel_app.Workbooks.Open(file_path)
    workbook.Save()
    workbook.Close(SaveChanges=True)
    excel_app.Quit()


def modify_excel_k1(file_path: str, k1_value: float, PL_value: float,
                    yco_value: float = None, bus_value: int = None):
    """解析条件 Excel の各セルにパラメータを書き込む。

    書き込み先セル:
        O1 = k1_value (ΔP または 継続時間)
        T4 = PL_value (負荷率)
        T5 = yco_value (短絡容量 [p.u.])  ※ 省略可
        T6 = bus_value (バス番号)          ※ 省略可
    """
    temp_file_path = file_path.replace('.xlsx', '_temp.xlsx')
    try:
        wb = load_workbook(file_path)
        sheet = wb.active
        sheet['O1'] = k1_value
        sheet['T4'] = PL_value
        if yco_value is not None:
            sheet['T5'] = yco_value
        if bus_value is not None:
            sheet['T6'] = bus_value
        wb.save(temp_file_path)
        wb.close()
        shutil.copy(temp_file_path, file_path)
        os.remove(temp_file_path)
        print("条件 Excel を更新しました。")
    except FileNotFoundError:
        print(f"[エラー] ファイルが見つかりません: {file_path}")
    except Exception as e:
        print(f"[エラー] Excel 更新中に例外が発生しました: {e}")


def operate_excel(file_path: str, mdl_base_name: str,
                  k1_value: float, ContingencyCase: int):
    """条件 Excel のセル O2 の値を読み取り、クリップボードにコピーして返す。"""
    try:
        wb = load_workbook(file_path)
        sheet = wb.active
        b3_value = math.floor(sheet["O2"].value)
        print(f"セル O2 の値: {b3_value}")
        wb.save(file_path)
        if ContingencyCase == 1:
            clipboard_string = f"00{mdl_base_name}-{int(k1_value)}MW-{b3_value}"
        elif ContingencyCase == 2:
            clipboard_string = f"00{mdl_base_name}-{int(k1_value*1000)}msec-{b3_value}"
        else:
            raise ValueError("ContingencyCase は 1 または 2 でなければなりません。")
        pyperclip.copy(clipboard_string)
        print(f"クリップボードにコピー: '{clipboard_string}'")
        return b3_value
    except Exception as e:
        print(f"[エラー] Excel 読み取り中に例外が発生しました: {e}")
        return None


def save_to_existing_excel(df: pd.DataFrame, file_path: str,
                            sheet_name: str, start_row: int, start_col: int):
    """既存の Excel ファイルの指定セルから DataFrame を書き込む。"""
    book = load_workbook(file_path)
    sheet = book[sheet_name]
    for i, row in df.iterrows():
        for j, value in enumerate(row):
            sheet.cell(row=start_row + i + 1, column=start_col + j + 1, value=value)
    book.save(file_path)


def execute_python_script(script_path: str):
    """指定した Python スクリプトをサブプロセスで実行する。"""
    result = subprocess.run(['python', script_path], capture_output=True, text=True)
    if result.returncode == 0:
        print("スクリプト実行成功。")
        print("出力:\n", result.stdout)
    else:
        print("[エラー] スクリプト実行失敗。")
        print("エラー出力:\n", result.stderr)


# =============================================================================
# GUI 操作ユーティリティ
# =============================================================================

def wait_for_window(title_keyword: str, timeout: float = 30) -> bool:
    """指定キーワードを含むウィンドウが現れるまで待機する。

    Args:
        title_keyword: ウィンドウタイトルに含まれるキーワード
        timeout: 最大待機時間 [秒]

    Returns:
        True: ウィンドウが見つかった / False: タイムアウト
    """
    deadline = time.time() + timeout
    while time.time() < deadline:
        wins = [w for w in gw.getAllWindows() if title_keyword in w.title]
        if wins:
            return True
        time.sleep(0.5)
    print(f"[警告] ウィンドウ '{title_keyword}' が {timeout} 秒以内に見つかりませんでした。")
    return False


def click_image(image_path: str, confidence: float = 0.8,
                wait: float = 0.5, timeout: float = 15) -> bool:
    """画面上の画像を探してクリックする（画像認識ベースのクリック）。

    Args:
        image_path: クリック対象の画像ファイルパス（PNG 推奨）
        confidence: 一致度のしきい値（0.0〜1.0）
        wait: クリック後の待機時間 [秒]
        timeout: 画像が見つかるまでの最大待機時間 [秒]

    Returns:
        True: クリック成功 / False: 画像が見つからなかった
    """
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            location = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if location:
                center = pyautogui.center(location)
                pyautogui.click(center)
                time.sleep(wait)
                return True
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(0.5)
    print(f"[警告] 画像が見つかりませんでした: {image_path}")
    return False


def replay_actions(file_name: str):
    """JSON マクロファイルを再生してマウス・キーボード操作を実行する。

    JSON マクロは座標ベースのため、画面解像度が異なると動作しない場合があります。
    可能であれば click_image() や send_shortcut() を使った操作に置き換えてください。
    """
    with open(file_name, 'r') as f:
        actions = json.load(f)

    start_time = actions[0][-1]
    mouse_controller = mouse.Controller()
    keyboard_controller = keyboard.Controller()

    for action in actions:
        action_type = action[0]
        if action_type == 'move':
            _, x, y, timestamp = action
            mouse_controller.position = (x, y)
        elif action_type == 'click':
            _, x, y, button, pressed, timestamp = action
            mouse_controller.position = (x, y)
            button = mouse.Button[button.split('.')[1]]
            if pressed:
                mouse_controller.press(button)
            else:
                mouse_controller.release(button)
        elif action_type == 'scroll':
            _, x, y, dx, dy, timestamp = action
            mouse_controller.scroll(dx, dy)
        elif action_type == 'key':
            _, key, event_type, timestamp = action
            try:
                if key.startswith('Key.'):
                    key = getattr(keyboard.Key, key.split('.')[1])
                if event_type == 'press':
                    keyboard_controller.press(key)
                else:
                    keyboard_controller.release(key)
            except AttributeError:
                print(f"[警告] 未対応のキー: {key}")
            except ValueError as e:
                print(f"[警告] キー解決エラー: {e}")

        time.sleep(max(0, timestamp - start_time))
        start_time = timestamp


def replay_all_actions(file_names: list, base_dir: str = None):
    """複数の JSON マクロファイルを順番に再生する。

    Args:
        file_names: JSON ファイル名のリスト
        base_dir: JSON ファイルが置かれているフォルダ（省略時は _BASE_DIR）
    """
    if base_dir is None:
        base_dir = _BASE_DIR
    for file_name in file_names:
        full_path = os.path.join(base_dir, file_name)
        print(f"マクロ再生中: {file_name}")
        replay_actions(full_path)
        print(f"マクロ再生完了: {file_name}")


def send_shortcut(*keys, wait: float = 0.3):
    """キーボードショートカットを送信する。

    Args:
        *keys: 送信するキー（例: 'ctrl', 's'）
        wait: 送信後の待機時間 [秒]

    使用例:
        send_shortcut('ctrl', 's')   # Ctrl+S（上書き保存）
        send_shortcut('alt', 'F4')   # Alt+F4（ウィンドウを閉じる）
    """
    pyautogui.hotkey(*keys)
    time.sleep(wait)


def maximize_window():
    """アクティブウィンドウを最大化する。"""
    window = gw.getActiveWindow()
    if window:
        window.maximize()
    else:
        print("[警告] アクティブなウィンドウが見つかりませんでした。")


def save_and_close_active_window():
    """アクティブウィンドウを保存して閉じる（Excel の場合は Ctrl+S を送信）。"""
    window = gw.getActiveWindow()
    if window:
        if 'Excel' in window.title:
            send_shortcut('ctrl', 's')
            window.close()
        else:
            window.close()
    else:
        print("[警告] アクティブなウィンドウが見つかりませんでした。")


# =============================================================================
# 組み合わせ生成
# =============================================================================

def generate_combinations(start: int, end: int, group_size: int, i: int) -> list:
    """ノード番号の重複組み合わせを生成し、group_size 単位のリストに分割する。"""
    numbers = range(start, end + 1)
    combinations = list(combinations_with_replacement(numbers, i))
    groups = [combinations[j:j + group_size]
              for j in range(0, len(combinations), group_size)]
    if len(groups[-1]) < group_size:
        needed = group_size - len(groups[-1])
        groups[-1].extend(groups[0][:needed])
    df_list = []
    for idx, group in enumerate(groups):
        df = pd.DataFrame(group, columns=[f'Element{k+1}' for k in range(i)])
        df_list.append(df)
    return df_list


def generate_combinations_all(res_list, M_list, yoc_list, bus_list, df_list, k1_values) -> list:
    """全パラメータの組み合わせリストを生成する。"""
    combinations = []
    for k1_value in k1_values:
        for res in res_list:
            for m_value in M_list:
                for yco_value in yoc_list:
                    for bus_value in bus_list:
                        for i, df in enumerate(df_list):
                            combinations.append((res, m_value, yco_value, bus_value, i, k1_value))
    return combinations


def create_combinations(TargetOfNumber: int, k1_start: float, k1_end: float,
                         delta_k1: float, df_list: list) -> list:
    """config.ini の設定に基づいて解析ケースの組み合わせを生成する。"""
    res_list = [0]
    M_list = [8]
    yoc_list = [0.25]
    bus_list = range(11, 48)

    if k1_start == 250:
        k1_values = list(range(int(k1_start), int(k1_end), int(delta_k1)))
    else:
        k1_values = [x / 100 for x in list(range(int(k1_start), int(k1_end), int(delta_k1)))]

    return generate_combinations_all(res_list, M_list, yoc_list, bus_list, df_list, k1_values)


# =============================================================================
# メイン解析ループ
# =============================================================================

def main(start_combination, combinations, sheet_name, start_row, start_col,
         PL_base, ContingencyCase, TargetOfNumber,
         folder_path, path_to_folder, condition_list_folder,
         script_path, excel_file_path, source_folder, source,
         destination_directory,
         original_file_sys, new_file_sys, new_directory_sys, new_zip_file_sys,
         original_file_condition, new_file_condition,
         new_directory_condition, new_zip_file_condition,
         Lp_i, Lq_i, df_list,
         window_wait_timeout, action_wait):
    """解析ケースを順番に実行するメインループ。"""

    file_name = os.path.basename(source)
    destination = os.path.join(destination_directory, file_name)

    # 条件テンプレートを condition0.xlsx としてコピー
    copy_and_rename_condition_file(ContingencyCase, TargetOfNumber,
                                   source_folder, folder_path)

    # 開始位置を特定
    try:
        start_index = combinations.index(start_combination) + 1
    except ValueError:
        start_index = 0

    remaining_combinations = combinations[start_index:]
    print(f"残りケース数: {len(remaining_combinations)}  "
          f"（推定時間: {len(remaining_combinations)*2/60:.1f} 時間）")

    for combination in remaining_combinations:
        res, m_value, yco_value, bus_value, i, k1_value = combination
        replacement_text = f"<Mg>{m_value}</Mg>"

        # 負荷・RES 値の計算
        value_pgo = calculate_value(Lp_i, res)
        value_qgo = calculate_value(Lq_i, res)
        value_text_pgo = f"<Pgo>{value_pgo}</Pgo>"
        value_text_qgo = f"<Qgo>{value_qgo}</Qgo>"

        Pgo = np.array([5.97560976, 9.3902439, 5.12195122, 9.3902439, 5.12195122,
                        9.3902439, 9.3902439, 5.97560976, 5.97560976, 4.26829268])
        Qgo_temp = np.array([2.47261939, 3.88554476, 2.11938805, 3.88554476, 2.11938805,
                              3.88554476, 3.88554476, 2.47261939, 2.47261939, 1.76615671])
        Qgo = np.array([2.66181, 3.86776, 1.45165, 3.7865, 1.6436,
                        3.97747, 5.75334, 1.73422, 1.82973, 2.25889])

        Pgo_total = np.sum(Pgo)
        Qgo_total = np.sum(Qgo)
        Pgo_percentages = Pgo / Pgo_total
        Pgo_res = calculate_value(Lp_i, res) * 37
        Qgo_res = calculate_value(Lq_i, res) * 37
        Pgo_res = Pgo - (Pgo_res * Pgo_percentages)
        Qgo_res = (Qgo_total - Qgo_res) * Pgo_percentages
        PL_value = PL_base * (1 - (res / 100))

        # 条件 Excel を更新
        if ContingencyCase == 1:
            modify_excel_k1(excel_file_path, k1_value / TargetOfNumber,
                            PL_value, yco_value, bus_value)
        elif ContingencyCase == 2:
            modify_excel_k1(excel_file_path, k1_value, PL_value)

        df = df_list[i]

        # CPAT を起動（JSON マクロ再生）
        replay_all_actions(["OPEN_CPAT.json"])
        # CPAT ウィンドウが現れるまで待機（堅牢化: タイムアウト付き）
        wait_for_window("CPAT", timeout=window_wait_timeout)

        # 条件を Excel に書き込み → VBA 実行 → MakeFile.py 実行
        save_to_existing_excel(df, excel_file_path, sheet_name, start_row, start_col)
        open_save_close_excel(excel_file_path)
        execute_python_script(script_path)
        b3_value = i

        # 系統モデルをコピーして解凍・編集・再圧縮
        shutil.copy2(source, destination)
        rename_files(original_file_sys, new_file_sys)
        rename_files(original_file_condition, new_file_condition)
        create_directory(new_directory_sys)
        create_directory(new_directory_condition)
        unzip_file(new_file_sys, new_directory_sys)
        unzip_file(new_file_condition, new_directory_condition)
        delete_files(new_directory_sys, ".pnsy")
        move_files(new_directory_condition, new_directory_sys, ".pnsy")
        replace_content_in_files(new_directory_sys, "<Mg>2.0</Mg>", replacement_text, ".pnsd")

        file_path_demand = os.path.join(new_directory_sys, "Demand.pnsj")
        modify_demand_file(file_path_demand, value_text_pgo, value_text_qgo)
        replace_tags_in_file(file_path_demand, Pgo, Pgo_res, Qgo_temp, Qgo_res)
        modify_yco_value(file_path_demand, bus_value, yco_value)

        compress_directory(new_directory_sys, new_zip_file_sys)
        rename_files(new_zip_file_sys, original_file_sys)
        clean_up(new_file_sys, new_file_condition, new_directory_sys, new_directory_condition)

        # 結果フォルダ名の生成
        mdl_base_name = (f"M{m_value}-RES{res*100}%-YCbus{int(bus_value)}"
                         f"-{int(yco_value * 1000)}MVA-ΔP{int(k1_value)}MW")
        if ContingencyCase == 1:
            clipboard_string = f"00{mdl_base_name}-{int(k1_value)}MW-{b3_value}"
        elif ContingencyCase == 2:
            clipboard_string = f"00{mdl_base_name}-{int(k1_value*1000)}msec-{b3_value}"
        else:
            raise ValueError("ContingencyCase は 1 または 2 でなければなりません。")
        pyperclip.copy(clipboard_string)

        # CPAT でシミュレーション実行（JSON マクロ再生）
        replay_all_actions(["DoCPAT.json"])

        # CPAT と Excel を強制終了
        for pid in get_process_ids('CPAT.exe'):
            kill_process_v1(pid)
        for pid in get_process_ids('EXCEL.EXE'):
            kill_process_v1(pid)

        # 後片付け
        delete_file(os.path.join(folder_path,
                                  "EAST10peak_AllThermal_AllDemand_auto_copy.pop"))
        rename_folder(
            os.path.join(path_to_folder, "EAST10peak_AllThermal_AllDemand_auto_res_SPM"),
            os.path.join(path_to_folder, str(clipboard_string))
        )
        os.remove(destination)

        time.sleep(action_wait)


# =============================================================================
# エントリーポイント
# =============================================================================

if __name__ == "__main__":
    # ------------------------------------------------------------------
    # config.ini を読み込む
    # ------------------------------------------------------------------
    cfg = load_config()

    # --- パス設定（config.ini の [paths] セクション） ---
    folder_path          = _resolve(cfg.get("paths", "folder_path"))
    path_to_folder       = _resolve(cfg.get("paths", "path_to_folder"))
    condition_list_folder = _resolve(cfg.get("paths", "condition_list_folder"))
    script_path          = _resolve(cfg.get("paths", "script_path"))
    excel_file_path      = _resolve(cfg.get("paths", "excel_file_path"))
    source_folder        = _resolve(cfg.get("paths", "source_folder"))
    source               = _resolve(cfg.get("paths", "source"))
    destination_directory = _resolve(cfg.get("paths", "destination_directory"))

    # --- 解析条件（config.ini の [analysis] セクション） ---
    ContingencyCase  = cfg.getint("analysis", "contingency_case")
    TargetOfNumber   = cfg.getint("analysis", "target_of_number")
    PL_base          = cfg.getfloat("analysis", "PL_base")
    Lp_i             = cfg.getfloat("analysis", "Lp_i")
    Lq_i             = cfg.getfloat("analysis", "Lq_i")
    node_start       = cfg.getint("analysis", "node_start")
    node_end         = cfg.getint("analysis", "node_end")
    sheet_name       = cfg.get("analysis", "sheet_name")
    start_row        = cfg.getint("analysis", "start_row")
    start_col        = cfg.getint("analysis", "start_col")

    if ContingencyCase == 1:
        k1_start = cfg.getfloat("analysis", "k1_start_L")
        k1_end   = cfg.getfloat("analysis", "k1_end_L")
        delta_k1 = cfg.getfloat("analysis", "delta_k1_L")
    else:
        k1_start = cfg.getfloat("analysis", "k1_start_GOC")
        k1_end   = cfg.getfloat("analysis", "k1_end_GOC")
        delta_k1 = cfg.getfloat("analysis", "delta_k1_GOC")

    # --- GUI 設定（config.ini の [gui] セクション） ---
    window_wait_timeout = cfg.getfloat("gui", "window_wait_timeout")
    action_wait         = cfg.getfloat("gui", "action_wait")

    # --- 定数定義（ファイル名は変更不要） ---
    original_file_sys       = "EAST10peak_AllThermal_AllDemand_auto_res_SPM.pop"
    new_file_sys            = "EAST10peak_AllThermal_AllDemand_auto_res_SPM.zip"
    new_directory_sys       = "EAST10peak_AllThermal_AllDemand_auto_res_SPM_unzipped"
    new_zip_file_sys        = "EAST10peak_AllThermal_AllDemand_auto_res_SPM_new.zip"
    original_file_condition = "condition0.pop"
    new_file_condition      = "condition0.zip"
    new_directory_condition = "condition0_unzipped"
    new_zip_file_condition  = "condition0_new.zip"

    # --- 組み合わせ生成 ---
    df_list     = generate_combinations(node_start, node_end, 10, TargetOfNumber)
    combinations = create_combinations(TargetOfNumber, k1_start, k1_end, delta_k1, df_list)

    # --- 解析開始位置の指定 ---
    # 途中から再開する場合は start_combination を変更してください。
    # 最初から実行する場合は combinations[0] のままで OK です。
    start_combination = combinations[0]

    # --- メインループ実行 ---
    main(
        start_combination, combinations,
        sheet_name, start_row, start_col,
        PL_base, ContingencyCase, TargetOfNumber,
        folder_path, path_to_folder, condition_list_folder,
        script_path, excel_file_path, source_folder, source,
        destination_directory,
        original_file_sys, new_file_sys, new_directory_sys, new_zip_file_sys,
        original_file_condition, new_file_condition,
        new_directory_condition, new_zip_file_condition,
        Lp_i, Lq_i, df_list,
        window_wait_timeout, action_wait
    )
