import shutil
import os
import time
import json
from pynput import mouse, keyboard
import pygetwindow as gw
import pyautogui
import subprocess
import pyperclip
import math
from openpyxl import load_workbook
import pandas as pd
import win32com.client
from itertools import combinations_with_replacement
import zipfile
import numpy as np
import subprocess

# ファイルの名前を変更する関数
def rename_files(original_file, new_file):
    os.rename(original_file, new_file)

# ディレクトリを作成する関数
def create_directory(directory):
    os.makedirs(directory, exist_ok=True)

# ZIPファイルを解凍する関数
def unzip_file(zip_file, extract_to):
    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
        zip_ref.extractall(extract_to)

# ディレクトリ内の特定のファイルを削除する関数
def delete_files(directory, extension):
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith(extension):
                os.remove(os.path.join(root, file))

# ファイルを移動する関数
def move_files(source_directory, destination_directory, extension):
    for root, dirs, files in os.walk(source_directory):
        for file in files:
            if file.endswith(extension):
                shutil.move(os.path.join(root, file), destination_directory)

# ファイルの内容を置換する関数
def replace_content_in_files(directory, target, replacement, extension):
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            if file.endswith(extension):
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                modified_content = content.replace(target, replacement)
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(modified_content)

# valueを計算する関数
def calculate_value(factor, res):
    return factor * res

# Demand.pnsj ファイルの内容を置換する関数
def modify_demand_file(file_path, value_text_pgo, value_text_qgo):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    modified_content = content.replace("<Pgo>0.0</Pgo>", value_text_pgo)
    modified_content = modified_content.replace("<Qgo>0.0</Qgo>", value_text_qgo)
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(modified_content)

# <Pgo>タグと<Qgo>タグを新しい値で置き換える関数
def replace_tags_in_file(file_path, old_values_pgo, new_values_pgo, old_values_qgo_temp, new_values_qgo):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    for old_value, new_value in zip(old_values_pgo, new_values_pgo):
        old_value_text = f"<Pgo>{old_value}</Pgo>"
        new_value_text = f"<Pgo>{new_value}</Pgo>"
        content = content.replace(old_value_text, new_value_text)
    for old_value, new_value in zip(old_values_qgo_temp, new_values_qgo):
        old_value_text = f"<Qgo>{old_value}</Qgo>"
        new_value_text = f"<Qgo>{new_value}</Qgo>"
        content = content.replace(old_value_text, new_value_text)
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(content)

# Yco 値を変更する関数
def modify_yco_value(file_path, bus_value, yco_value):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    lines = content.split('\n')
    for i, line in enumerate(lines):
        if f"<Number>{bus_value}</Number>" in line:
            if i + 7 < len(lines) and "<Yco>" in lines[i + 7]:
                lines[i + 7] = f"          <Yco>{yco_value}</Yco>"
    with open(file_path, 'w', encoding='utf-8') as file:
        file.write('\n'.join(lines))
    print(f"{os.path.basename(file_path)} の <Number>{bus_value}</Number> の部分の <Yco> を {yco_value} に変更しました。")

# ディレクトリを圧縮する関数
def compress_directory(directory, output_file):
    shutil.make_archive(output_file.replace('.zip', ''), 'zip', directory)

# 中間ファイルと作業ディレクトリを削除する関数
def clean_up(*files_and_directories):
    for item in files_and_directories:
        if os.path.isfile(item):
            os.remove(item)
        elif os.path.isdir(item):
            shutil.rmtree(item)

# ファイルをコピー
def copy_file(folder_path, file_name):
    source_path = os.path.join(folder_path, file_name)
    destination_path = os.path.join(folder_path, file_name.replace(".pop", "_copy.pop"))
    shutil.copy2(source_path, destination_path)
    print(f"ファイル {file_name} を {destination_path} にコピーしました。")

def execute_python_script(script_path):
    # try:
    # サブプロセスを使ってスクリプトを実行
    result = subprocess.run(['python', script_path], capture_output=True, text=True)
    
    # 実行結果の表示
    if result.returncode == 0:
        print("Script executed successfully.")
        print("Output:\\n", result.stdout)
    else:
        print("Script execution failed.")
        print("Error:\\n", result.stderr)
    
    # except Exception as e:
    #     print(f"An error occurred: {e}")

# マウスとキーボードのアクションを再生
def replay_actions(file_name):
    with open(file_name, 'r') as f:
        actions = json.load(f)

    start_time = actions[0][-1]
    mouse_controller = mouse.Controller()
    keyboard_controller = keyboard.Controller()

    for action in actions:
        action_type = action[0]
        if action_type == 'move':
            _, x, y, timestamp = action
            mouse_controller.position = (x, y)
        elif action_type == 'click':
            _, x, y, button, pressed, timestamp = action
            mouse_controller.position = (x, y)
            button = mouse.Button[button.split('.')[1]]
            if pressed:
                mouse_controller.press(button)
            else:
                mouse_controller.release(button)
        elif action_type == 'scroll':
            _, x, y, dx, dy, timestamp = action
            mouse_controller.scroll(dx, dy)
        elif action_type == 'key':
            _, key, event_type, timestamp = action
            try:
                if key.startswith('Key.'):
                    key = getattr(keyboard.Key, key.split('.')[1])
                if event_type == 'press':
                    keyboard_controller.press(key)
                else:
                    keyboard_controller.release(key)
            except AttributeError:
                print(f"未対応のキー: {key}")
            except ValueError as e:
                print(f"キーの解決中にエラーが発生しました: {e}")

        time.sleep(max(0, timestamp - start_time))
        start_time = timestamp

def replay_all_actions(file_names):
    for file_name in file_names:
        print(f"再生中: {file_name}")
        replay_actions(file_name)
        print(f"再生完了: {file_name}")

# Excelファイルを操作
def operate_excel(file_path, mdl_base_name, k1_value, ContingencyCase):
    try:
        # openpyxlを使ってExcelファイルを読み込む
        from openpyxl import load_workbook
        import math
        import pyperclip

        workbook = load_workbook(file_path)
        sheet = workbook.active
        b3_value = math.floor(sheet["O2"].value)
        print(f"セルO2の値: {b3_value}")
        workbook.save(file_path)
        
        # ContingencyCaseに応じてclipboard_stringを変更
        if ContingencyCase == 1:
            clipboard_string = f"00{mdl_base_name}-{int(k1_value)}MW-{b3_value}"
        elif ContingencyCase == 2:
            clipboard_string = f"00{mdl_base_name}-{int(k1_value*1000)}msec-{b3_value}"
        else:
            raise ValueError("ContingencyCaseは1または2でなければなりません。")
        
        pyperclip.copy(clipboard_string)
        print(f"文字列 '{clipboard_string}' をクリップボードにコピーしました。")
        return b3_value
    except Exception as e:
        print(f"An error occurred: {e}")
        return None

# ウィンドウを最大化
def maximize_window():
    window = gw.getActiveWindow()
    if window:
        window.maximize()
    else:
        print("アクティブなウィンドウが見つかりませんでした。")

# アクティブウィンドウを保存して閉じる
def save_and_close_active_window():
    window = gw.getActiveWindow()
    if window:
        if 'Excel' in window.title:
            pyautogui.hotkey('ctrl', 's')
            time.sleep(1)
            window.close()
        else:
            window.close()
    else:
        print("アクティブなウィンドウが見つかりませんでした。")

# ファイルを削除
def delete_file(file_path):
    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"{file_path} を削除しました。")
    else:
        print(f"{file_path} が見つかりません。")

def delete_all_files_in_folder(folder_path):
    # フォルダが存在するか確認
    if os.path.exists(folder_path):
        # フォルダ内のすべてのファイルを削除
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)  # ファイルまたはリンクを削除
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)  # ディレクトリを削除
            except Exception as e:
                print(f'Failed to delete {file_path}. Reason: {e}')
    else:
        print(f'The folder {folder_path} does not exist.')

def rename_folder(old_folder_name, new_folder_name):
    try:
        os.rename(old_folder_name, new_folder_name)
        print(f"フォルダ名が '{old_folder_name}' から '{new_folder_name}' に変更されました。")
    except Exception as e:
        print(f"フォルダ名の変更中にエラーが発生しました: {e}")

def open_save_close_excel(file_path):
    # try:
    # Excelアプリケーションを起動
    excel_app = win32com.client.Dispatch("Excel.Application")
    excel_app.Visible = False  # Excelを表示しない
    excel_app.DisplayAlerts = False  # 警告を表示しない
    
    # ワークブックを開く
    workbook = excel_app.Workbooks.Open(file_path)
    
    # 変更を保存
    workbook.Save()
    workbook.Close(SaveChanges=True)
    excel_app.Quit()
        
    #     print("File opened, VBA executed, saved, and closed successfully.")
        
    # except Exception as e:
    #     print(f"An error occurred: {e}")

def modify_excel_k1(file_path, k1_value, PL_value, yco_value, bus_value):
    try:
        # 一時ファイルのパスを作成
        temp_file_path = file_path.replace('.xlsx', '_temp.xlsx')
        
        # openpyxlを使用してExcelファイルを操作
        wb = load_workbook(file_path)
        sheet = wb.active
        sheet['O1'] = k1_value
        sheet['T4'] = PL_value
        sheet['T5'] = yco_value
        sheet['T6'] = bus_value
        
        # 一時ファイルとして保存
        wb.save(temp_file_path)
        wb.close()
        
        # 一時ファイルを元のファイルに上書きコピー
        shutil.copy(temp_file_path, file_path)
        os.remove(temp_file_path)  # 一時ファイルを削除
        
        print("File opened, K1 value set, saved, and closed successfully.")
        
    except FileNotFoundError:
        print(file_path)
        print("File not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

def generate_combinations(start, end, group_size, i):
    # 数字の範囲を定義
    numbers = range(start, end + 1)

    # 重複組み合わせを生成
    combinations = list(combinations_with_replacement(numbers, i))

    # 組み合わせをグループに分ける
    groups = [combinations[j:j + group_size] for j in range(0, len(combinations), group_size)]

    # 最後のグループを補完するために、最初のグループから必要な数を追加
    if len(groups[-1]) < group_size:
        needed = group_size - len(groups[-1])
        groups[-1].extend(groups[0][:needed])

    # 各グループをDataFrameに変換
    df_list = []
    for idx, group in enumerate(groups):
        df = pd.DataFrame(group, columns=[f'Element{k+1}' for k in range(i)])
        df_list.append(df)

    return df_list

def save_to_existing_excel(df, file_path, sheet_name, start_row, start_col):
    # 既存のエクセルファイルを読み込む
    book = load_workbook(file_path)
    sheet = book[sheet_name]

    # データを書き込む
    for i, row in df.iterrows():
        for j, value in enumerate(row):
            sheet.cell(row=start_row + i + 1, column=start_col + j + 1, value=value)

    # ファイルを保存
    book.save(file_path)

def copy_and_rename_condition_file(contingency_case, target_of_number, source_folder, destination_folder):
    if contingency_case == 1 and target_of_number == 1:
        src_file = "Condition(L)_1site.xlsx"
    elif contingency_case == 1 and target_of_number == 2:
        src_file = "Condition(L)_2site.xlsx"
    elif contingency_case == 2 and target_of_number == 1:
        src_file = "Condition(G_O_C)_1site.xlsx"
    elif contingency_case == 2 and target_of_number == 2:
        src_file = "Condition(G_O_C)_2site.xlsx"
    else:
        raise ValueError("Invalid combination of ContingencyCase and TargetOfNumber")
    
    src_path = os.path.join(source_folder, src_file)
    dst_path = os.path.join(destination_folder, "condition0.xlsx")
    
    if os.path.exists(src_path):
        shutil.copy(src_path, dst_path)
    else:
        raise FileNotFoundError(f"Source file {src_path} does not exist")

def copy_files2(main_folder, condition_list_folder, src_file_name, dst_file_name):
    src_mdl_path = os.path.join(condition_list_folder, src_file_name)
    dst_main_path = os.path.join(main_folder, dst_file_name)
    backup_mdl_path = os.path.join(condition_list_folder, dst_file_name)

    if os.path.exists(src_mdl_path):
        # 元のEAST10peak_AllThermal_AllDemand_auto.popをバックアップフォルダに移動
        if os.path.exists(dst_main_path):
            shutil.move(dst_main_path, backup_mdl_path)
        
        # MDLファイルをmainフォルダにコピー
        shutil.copy(src_mdl_path, dst_main_path)
        
        # ファイル名をEAST10peak_AllThermal_AllDemand_auto.popに変更
        os.rename(dst_main_path, os.path.join(main_folder, dst_file_name))

def get_process_ids(process_name):
    result = subprocess.run(['tasklist'], capture_output=True, text=True)
    lines = result.stdout.splitlines()
    pids = []

    for line in lines:
        if process_name.lower() in line.lower():
            parts = line.split()
            pid = int(parts[1])
            pids.append(pid)
    
    return pids

def kill_process(pid):
    subprocess.run(['taskkill', '/PID', str(pid), '/F'])

# メイン処理
if __name__ == "__main__":
    # 諸々設定
    sheet_name = 'Sheet1'  # 書き込みたいシート名
    start_row = 1
    start_col = 11
    PL_base = 1.89189189
    ContingencyCase = 1 # 1: 負荷増加, 2: 短絡⇒遮断⇒閉路
    TargetOfNumber = 1  # 事故箇所数
    res_list = [0.2, 0.5, 0.8] # res_list = [0.2, 0.5, 0.8]
    M_list = [2, 8, 16] # M_list = [2, 8, 16]
    yoc_list = [0.5]
    bus_list = range(11, 48)

    if ContingencyCase == 1:
        k1_value = 250
        k1_increment = 250
        k1_max = 1000
        # 重複組み合わせを生成
        df_list = generate_combinations(11, 47, 10, TargetOfNumber)
    elif ContingencyCase == 2:
        k1_value = 0.15
        k1_increment = 0.01
        k1_max = 0.15
    
    folder_path = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\02_MyModel\\CPATFreeR05V1.0\\CPATFree_R05_Ver1.0\\model\\EAST10\\main"
    file_name = "EAST10peak_AllThermal_AllDemand_auto.pop"
    path_to_folder = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\02_MyModel\\CPATFreeR05V1.0\\CPATFree_R05_Ver1.0\\Temp"
    condition_list_folder = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\02_MyModel\\CPATFreeR05V1.0\\CPATFree_R05_Ver1.0\\model\\EAST10\\List\\MDL"
    script_path = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\02_MyModel\\CPATFreeR05V1.0\\CPATFree_R05_Ver1.0\\model\\EAST10\\main\\MakeFile.py"
    excel_file_path = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\02_MyModel\\CPATFreeR05V1.0\\CPATFree_R05_Ver1.0\\model\\EAST10\\main\\condition0.xlsx"
    source_folder = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\02_MyModel\\CPATFreeR05V1.0\\CPATFree_R05_Ver1.0\\model\\EAST10\\List\\Condition"
    source = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\02_MyModel\\CPATFreeR05V1.0\\CPATFree_R05_Ver1.0\\model\\EAST10\\List\\MDL\\EAST10peak_AllThermal_AllDemand_auto_res_SPM.pop"
    destination_directory = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\02_MyModel\\CPATFreeR05V1.0\\CPATFree_R05_Ver1.0\\model\\EAST10\\main"

    # 定数定義
    original_file_sys = "EAST10peak_AllThermal_AllDemand_auto_res_SPM.pop"
    new_file_sys = "EAST10peak_AllThermal_AllDemand_auto_res_SPM.zip"
    new_directory_sys = "EAST10peak_AllThermal_AllDemand_auto_res_SPM_unzipped"
    new_zip_file_sys = "EAST10peak_AllThermal_AllDemand_auto_res_SPM_new.zip"

    original_file_condition = "condition0.pop"
    new_file_condition = "condition0.zip"
    new_directory_condition = "condition0_unzipped"
    new_zip_file_condition = "condition0_new.zip"

    # ファイル名
    file_name = os.path.basename(source)

    # コピー先のファイルパス
    destination = os.path.join(destination_directory, file_name)

    # Excelファイルのコピーと名前変更
    copy_and_rename_condition_file(ContingencyCase, TargetOfNumber, source_folder, folder_path)

    while k1_value <= k1_max:
        for res in res_list:
            for m_value in M_list:

                replacement_text = f"<Mg>{m_value}</Mg>"

                Lp_i = 1.89189189
                Lq_i = 0.66824879

                value_pgo = calculate_value(Lp_i, res)
                value_qgo = calculate_value(Lq_i, res)
                value_text_pgo = f"<Pgo>{value_pgo}</Pgo>"
                value_text_qgo = f"<Qgo>{value_qgo}</Qgo>"

                Pgo = np.array([5.97560976, 9.3902439, 5.12195122, 9.3902439, 5.12195122, 9.3902439, 9.3902439, 5.97560976, 5.97560976, 4.26829268])
                Qgo_temp = np.array([2.47261939, 3.88554476, 2.11938805, 3.88554476, 2.11938805, 3.88554476, 3.88554476, 2.47261939, 2.47261939, 1.76615671])
                Qgo = np.array([2.66181, 3.86776, 1.45165, 3.7865, 1.6436, 3.97747, 5.75334, 1.73422, 1.82973, 2.25889])

                Pgo_total = np.sum(Pgo)
                Qgo_total = np.sum(Qgo)

                Pgo_percentages = (Pgo / Pgo_total)

                Pgo_res = calculate_value(Lp_i, res) * 37
                Qgo_res = calculate_value(Lq_i, res) * 37

                Pgo_res = Pgo - (Pgo_res * Pgo_percentages)
                Qgo_res = (Qgo_total - Qgo_res) * Pgo_percentages

                for yco_value in yoc_list:
                    for bus_value in bus_list:
                        # mdl_files = [f"M8-Res{res}.pop", f"M2-Res{res}.pop", f"M16-Res{res}.pop"]
                        PL_value = PL_base * ( 1 - (res / 100))

                        # Excelファイルの値更新
                        if ContingencyCase == 1:
                            modify_excel_k1(excel_file_path, k1_value / TargetOfNumber, PL_value, yco_value, bus_value)
                        elif ContingencyCase == 2:
                            modify_excel_k1(excel_file_path, k1_value, PL_value)
                    
                        # copy_files2(folder_path, condition_list_folder, mdl_file, file_name)
                        # mdl_base_name = os.path.splitext(mdl_file)[0]
                        
                        for i, df in enumerate(df_list):
                            # i = 1
                            # df = df_list[i]

                            action_files = ["OpenCPAT.json"]
                            replay_all_actions(action_files)
                            maximize_window()

                            # 既存のエクセルファイルにデータを書き込む
                            save_to_existing_excel(df, excel_file_path, sheet_name, start_row, start_col)
                            
                            # copy_file(folder_path, file_name)

                            # 条件設定
                            open_save_close_excel(excel_file_path)
                            execute_python_script(script_path)
                            b3_value = i

                            # 系統コピー
                            shutil.copy2(source, destination)

                            rename_files(original_file_sys, new_file_sys)
                            rename_files(original_file_condition, new_file_condition)

                            # 新しいディレクトリを作成
                            create_directory(new_directory_sys)
                            create_directory(new_directory_condition)

                            # ZIPファイルを解凍
                            unzip_file(new_file_sys, new_directory_sys)
                            unzip_file(new_file_condition, new_directory_condition)

                            # ディレクトリ内の .pnsy ファイルを削除
                            delete_files(new_directory_sys, ".pnsy")

                            # condition0_unzipped 内の .pnsy ファイルを EAST10peak_AllThermal_AllDemand_auto_res_SPM_unzipped に移動
                            move_files(new_directory_condition, new_directory_sys, ".pnsy")

                            # 解凍されたファイルを処理
                            replace_content_in_files(new_directory_sys, "<Mg>2.0</Mg>", replacement_text, ".pnsd")

                            # Demand.pnsj ファイルを変更
                            file_path = os.path.join(new_directory_sys, "Demand.pnsj")
                            modify_demand_file(file_path, value_text_pgo, value_text_qgo)

                            # <Pgo>タグと<Qgo>タグを新しい値で置き換え
                            replace_tags_in_file(file_path, Pgo, Pgo_res, Qgo_temp, Qgo_res)

                            # Yco 値を変更
                            modify_yco_value(file_path, bus_value, yco_value)

                            # ディレクトリを圧縮
                            compress_directory(new_directory_sys, new_zip_file_sys)

                            # 圧縮したファイルの名前を元に戻す
                            rename_files(new_zip_file_sys, original_file_sys)

                            # 作業ディレクトリと中間ファイルを削除
                            clean_up(new_file_sys, new_file_condition, new_directory_sys, new_directory_condition)

                            print("処理が完了しました。")

                            # ContingencyCaseに応じてclipboard_stringを変更
                            if ContingencyCase == 1:
                                clipboard_string = f"M{m_value}-RES{res*100}%-YCbus{int(bus_value)}-{int(yco_value * 1000)}MVA-ΔP{int(k1_value)}MW-{b3_value}"
                            elif ContingencyCase == 2:
                                clipboard_string = f"00{mdl_base_name}-{int(k1_value*1000)}msec-{b3_value}"
                            else:
                                raise ValueError("ContingencyCaseは1または2でなければなりません。")
                            pyperclip.copy(clipboard_string)
                          
                            # # b3_value = operate_excel(excel_file_path, mdl_base_name, k1_value, ContingencyCase)

                            # if b3_value is None:
                            #     continue

                            # action_files = ["ExecuteCPAT.json"]
                            # replay_all_actions(action_files)
                            
                            action_files = ["DoCPAT.json"]
                            replay_all_actions(action_files)
                            
                            # save_and_close_active_window()
                            # time.sleep(1)

                            # CPATとExcelのプロセスIDを取得し，終了
                            cpat_pids = get_process_ids('CPAT.exe')
                            excel_pids = get_process_ids('EXCEL.EXE')
                            for pid in cpat_pids:
                                kill_process(pid)
                            
                            for pid in excel_pids:
                                kill_process(pid)

                            # action_files = ["CloseCPAT.json"]
                            # replay_all_actions(action_files)

                            delete_file(os.path.join(folder_path, "EAST10peak_AllThermal_AllDemand_auto_copy.pop"))
                            # delete_all_files_in_folder(os.path.join(path_to_folder, f"EAST10peak_AllThermal_AllDemand_auto_copy_{b3_value}"))
                            rename_folder(os.path.join(path_to_folder, "EAST10peak_AllThermal_AllDemand_auto_res_SPM"), os.path.join(path_to_folder, str(clipboard_string)))
                            os.remove(destination)
                    
        k1_value += k1_increment
