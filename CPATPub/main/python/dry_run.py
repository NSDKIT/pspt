# =============================================================================
# CPAT 自動解析システム ドライラン（GUI なし代替シミュレーション）
# =============================================================================
# CPAT-GUI が使えない環境（Linux など）で、以下を検証します。
#   1. config.ini の読み込み
#   2. 解析ケースの組み合わせ生成
#   3. condition0.xlsx への条件書き込み
#   4. .pop ファイルの展開 → パラメータ書き換え → 再圧縮
# =============================================================================

import os
import sys
import shutil
import zipfile
import math
import configparser
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from itertools import combinations_with_replacement

# ---- パス解決ヘルパー -------------------------------------------------------
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def _resolve(path: str) -> str:
    if os.path.isabs(path):
        return path
    return os.path.normpath(os.path.join(_BASE_DIR, path))

# ---- config.ini 読み込み ----------------------------------------------------
def load_config():
    cfg = configparser.ConfigParser()
    cfg_path = os.path.join(_BASE_DIR, "config.ini")
    cfg.read(cfg_path, encoding="utf-8")
    return cfg

# ---- 組み合わせ生成（main_v4.py から抜粋） -----------------------------------
def generate_combinations(start, end, group_size, i):
    buses = list(range(start, end + 1))
    return list(combinations_with_replacement(buses, i))

def create_combinations(TargetOfNumber, k1_start, k1_end, delta_k1):
    res_list   = [0]
    M_list     = [8]
    yco_list   = [0.25]
    bus_list   = range(11, 48)

    if k1_start == 250:
        k1_values = list(range(int(k1_start), int(k1_end), int(delta_k1)))
    else:
        k1_values = [x / 100 for x in range(int(k1_start), int(k1_end), int(delta_k1))]

    df_list = generate_combinations(11, 47, 10, TargetOfNumber)

    combos = []
    for k1 in k1_values:
        for res in res_list:
            for m in M_list:
                for yco in yco_list:
                    for bus in bus_list:
                        for i, df in enumerate(df_list):
                            combos.append((res, m, yco, bus, i, k1))
    return combos, k1_values, df_list

# ---- Excel 書き込み（main_v4.py の modify_excel_k1 から抜粋） ---------------
def modify_excel_k1(file_path, k1_value, PL_value, yco_value=None, bus_value=None):
    temp_path = file_path.replace('.xlsx', '_temp.xlsx')
    wb = load_workbook(file_path)
    ws = wb.active
    ws['O1'] = k1_value
    ws['T4'] = PL_value
    if yco_value is not None:
        ws['T5'] = yco_value
    if bus_value is not None:
        ws['T6'] = bus_value
    wb.save(temp_path)
    wb.close()
    shutil.copy(temp_path, file_path)
    os.remove(temp_path)

# ---- .pop ファイルのパラメータ書き換え ---------------------------------------
def modify_pop_params(source_pop, dest_pop, pgo_new, qgo_new, yco_value, bus_value):
    """
    .pop（ZIP）を展開し、pf.pnsj の Pgo/Qgo と Yco を書き換えて再圧縮する。
    """
    import subprocess
    work_dir = dest_pop.replace('.pop', '_work')
    os.makedirs(work_dir, exist_ok=True)

    # 展開（unzip コマンドで Shift-JIS ファイル名を迷わず展開）
    subprocess.run(
        ['unzip', '-o', '-j', source_pop, '-d', work_dir],
        capture_output=True
    )

    # pf.pnsj を書き換え
    pnsj_path = os.path.join(work_dir, 'pf.pnsj')
    tree = ET.parse(pnsj_path)
    root = tree.getroot()

    gen_nodes = [n for n in root.findall('.//SNode') if n.findtext('G') == 'true']
    total_gen = len(gen_nodes)
    pgo_per_gen = pgo_new / total_gen if total_gen > 0 else 0
    qgo_per_gen = qgo_new / total_gen if total_gen > 0 else 0

    for n in gen_nodes:
        pgo_el = n.find('Pgo')
        qgo_el = n.find('Qgo')
        if pgo_el is not None:
            pgo_el.text = f"{pgo_per_gen:.8f}"
        if qgo_el is not None:
            qgo_el.text = f"{qgo_per_gen:.8f}"

    # 指定バスの Yco を書き換え
    for n in root.findall('.//SNode'):
        if n.findtext('Number') == str(bus_value):
            yco_el = n.find('Yco')
            if yco_el is not None:
                yco_el.text = str(yco_value)

    tree.write(pnsj_path, encoding='utf-8', xml_declaration=True)

    # 再圧縮
    with zipfile.ZipFile(dest_pop, 'w', zipfile.ZIP_DEFLATED) as zf:
        for root_dir, dirs, files in os.walk(work_dir):
            for file in files:
                file_path = os.path.join(root_dir, file)
                arcname = os.path.relpath(file_path, work_dir)
                zf.write(file_path, arcname)

    shutil.rmtree(work_dir)
    return pgo_per_gen, qgo_per_gen

# =============================================================================
# メイン処理
# =============================================================================
if __name__ == "__main__":
    print("=" * 60)
    print("  CPAT 自動解析システム ドライラン（GUI なし代替）")
    print("=" * 60)

    # --- config.ini 読み込み ---
    cfg = load_config()
    ContingencyCase  = cfg.getint("analysis", "contingency_case")
    TargetOfNumber   = cfg.getint("analysis", "target_of_number")
    PL_base          = cfg.getfloat("analysis", "PL_base")
    node_start       = cfg.getint("analysis", "node_start")
    node_end         = cfg.getint("analysis", "node_end")
    source_pop       = _resolve(cfg.get("paths", "source"))
    excel_file_path  = _resolve(cfg.get("paths", "excel_file_path"))
    path_to_folder   = _resolve(cfg.get("paths", "path_to_folder"))

    if ContingencyCase == 1:
        k1_start = cfg.getfloat("analysis", "k1_start_L")
        k1_end   = cfg.getfloat("analysis", "k1_end_L")
        delta_k1 = cfg.getfloat("analysis", "delta_k1_L")
        case_label = "負荷増加 (ΔP [MW])"
    else:
        k1_start = cfg.getfloat("analysis", "k1_start_GOC")
        k1_end   = cfg.getfloat("analysis", "k1_end_GOC")
        delta_k1 = cfg.getfloat("analysis", "delta_k1_GOC")
        case_label = "短絡継続時間 [サイクル]"

    print(f"\n[設定] 事故種別: {ContingencyCase} ({case_label})")
    print(f"[設定] 事故箇所数: {TargetOfNumber}")
    print(f"[設定] k1 範囲: {k1_start} ～ {k1_end - delta_k1} (刻み {delta_k1})")
    print(f"[設定] ノード範囲: {node_start} ～ {node_end}")
    print(f"[設定] ベースモデル: {source_pop}")

    # --- Step 1: 組み合わせ生成 ---
    print("\n" + "-" * 60)
    print("Step 1: 解析ケースの組み合わせを生成します...")
    combos, k1_values, df_list = create_combinations(
        TargetOfNumber, k1_start, k1_end, delta_k1
    )
    print(f"  k1 の値リスト: {k1_values}")
    print(f"  バス組み合わせ数: {len(df_list)}")
    print(f"  総解析ケース数: {len(combos)}")
    print(f"  先頭5ケース: {combos[:5]}")

    # --- Step 2: Excel への条件書き込み（最初の3ケース分） ---
    print("\n" + "-" * 60)
    print("Step 2: Excel への条件書き込みをテストします（先頭3ケース）...")
    os.makedirs(path_to_folder, exist_ok=True)

    results_log = []
    for idx, combo in enumerate(combos[:3]):
        res, m_value, yco_value, bus_value, df_idx, k1_value = combo
        PL_value = PL_base * (1 + res)
        modify_excel_k1(excel_file_path, k1_value, PL_value, yco_value, bus_value)
        wb = load_workbook(excel_file_path)
        ws = wb.active
        o1 = ws['O1'].value
        t4 = ws['T4'].value
        t5 = ws['T5'].value
        t6 = ws['T6'].value
        wb.close()
        print(f"  ケース {idx+1}: k1={k1_value}, PL={PL_value:.4f}, "
              f"Yco={yco_value}, Bus={bus_value} → Excel O1={o1}, T4={t4:.4f}, T5={t5}, T6={t6}")
        results_log.append({
            "ケース番号": idx + 1,
            "k1": k1_value,
            "PL_value": round(PL_value, 4),
            "Yco": yco_value,
            "バス番号": bus_value,
            "Excel_O1": o1,
            "Excel_T4": round(t4, 4),
            "Excel_T5": t5,
            "Excel_T6": t6,
        })

    # --- Step 3: .pop ファイルのパラメータ書き換え（1ケース分） ---
    print("\n" + "-" * 60)
    print("Step 3: .pop ファイルのパラメータ書き換えをテストします（1ケース分）...")

    if not os.path.exists(source_pop):
        # フォールバック: mdl/ にある実際のファイルを使う
        fallback = os.path.join(_BASE_DIR, "../mdl/275-500kVloop_EAST10peak_10SG_3RES_1SC.pop")
        source_pop = _resolve(fallback)
        print(f"  [注意] config.ini の source が見つからないため代替ファイルを使用: {source_pop}")

    dest_pop = os.path.join(path_to_folder, "test_case_001.pop")
    combo = combos[0]
    res, m_value, yco_value, bus_value, df_idx, k1_value = combo
    pgo_new = 0.35 * (1 - res)
    qgo_new = 0.12 * (1 - res)

    pgo_per, qgo_per = modify_pop_params(
        source_pop, dest_pop, pgo_new, qgo_new, yco_value, bus_value
    )
    print(f"  書き換え内容: Pgo/発電機={pgo_per:.8f}, Qgo/発電機={qgo_per:.8f}, "
          f"Yco={yco_value} (バス{bus_value})")
    print(f"  出力ファイル: {dest_pop}")

    # 書き換え後の値を検証
    with zipfile.ZipFile(dest_pop, 'r') as zf:
        with zf.open('pf.pnsj') as f:
            tree = ET.parse(f)
            root = tree.getroot()
    gen_nodes = [n for n in root.findall('.//SNode') if n.findtext('G') == 'true']
    print(f"  検証: 発電機数={len(gen_nodes)}, "
          f"G1の Pgo={gen_nodes[0].findtext('Pgo')}")

    # --- 結果サマリー ---
    print("\n" + "=" * 60)
    print("  ドライラン完了！")
    print("=" * 60)
    df_result = pd.DataFrame(results_log)
    print("\n[Excel 書き込み結果サマリー]")
    print(df_result.to_string(index=False))

    # CSV に保存
    csv_out = os.path.join(path_to_folder, "dry_run_results.csv")
    df_result.to_csv(csv_out, index=False, encoding="utf-8-sig")
    print(f"\n結果を CSV に保存しました: {csv_out}")
