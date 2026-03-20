import shutil
import os
import time
import json
from pynput import mouse, keyboard
import pygetwindow as gw
import pyautogui
import subprocess
import pyperclip
import math
from openpyxl import load_workbook
import pandas as pd
import win32com.client

# ファイルをコピー
def copy_file(folder_path, file_name):
    source_path = os.path.join(folder_path, file_name)
    destination_path = os.path.join(folder_path, file_name.replace(".pop", "_copy.pop"))
    shutil.copy2(source_path, destination_path)
    print(f"ファイル {file_name} を {destination_path} にコピーしました。")

def execute_python_script(script_path):
    # try:
    # サブプロセスを使ってスクリプトを実行
    result = subprocess.run(['python', script_path], capture_output=True, text=True)
    
    # 実行結果の表示
    if result.returncode == 0:
        print("Script executed successfully.")
        print("Output:\\n", result.stdout)
    else:
        print("Script execution failed.")
        print("Error:\\n", result.stderr)
    
    # except Exception as e:
    #     print(f"An error occurred: {e}")

# マウスとキーボードのアクションを再生
def replay_actions(file_name):
    with open(file_name, 'r') as f:
        actions = json.load(f)

    start_time = actions[0][-1]
    mouse_controller = mouse.Controller()
    keyboard_controller = keyboard.Controller()

    for action in actions:
        action_type = action[0]
        if action_type == 'move':
            _, x, y, timestamp = action
            mouse_controller.position = (x, y)
        elif action_type == 'click':
            _, x, y, button, pressed, timestamp = action
            mouse_controller.position = (x, y)
            button = mouse.Button[button.split('.')[1]]
            if pressed:
                mouse_controller.press(button)
            else:
                mouse_controller.release(button)
        elif action_type == 'scroll':
            _, x, y, dx, dy, timestamp = action
            mouse_controller.scroll(dx, dy)
        elif action_type == 'key':
            _, key, event_type, timestamp = action
            try:
                if key.startswith('Key.'):
                    key = getattr(keyboard.Key, key.split('.')[1])
                if event_type == 'press':
                    keyboard_controller.press(key)
                else:
                    keyboard_controller.release(key)
            except AttributeError:
                print(f"未対応のキー: {key}")
            except ValueError as e:
                print(f"キーの解決中にエラーが発生しました: {e}")

        time.sleep(max(0, timestamp - start_time))
        start_time = timestamp

def replay_all_actions(file_names):
    for file_name in file_names:
        print(f"再生中: {file_name}")
        replay_actions(file_name)
        print(f"再生完了: {file_name}")

# Excelファイルを操作
def operate_excel(file_path, mdl_base_name, k1_value, ContingencyCase):
    try:
        # openpyxlを使ってExcelファイルを読み込む
        from openpyxl import load_workbook
        import math
        import pyperclip

        workbook = load_workbook(file_path)
        sheet = workbook.active
        b3_value = math.floor(sheet["O2"].value)
        print(f"セルO2の値: {b3_value}")
        workbook.save(file_path)
        
        # ContingencyCaseに応じてclipboard_stringを変更
        if ContingencyCase == 1:
            clipboard_string = f"00{mdl_base_name}-{int(k1_value)}MW-{b3_value}"
        elif ContingencyCase == 2:
            clipboard_string = f"00{mdl_base_name}-{int(k1_value*1000)}msec-{b3_value}"
        else:
            raise ValueError("ContingencyCaseは1または2でなければなりません。")
        
        pyperclip.copy(clipboard_string)
        print(f"文字列 '{clipboard_string}' をクリップボードにコピーしました。")
        return b3_value
    except Exception as e:
        print(f"An error occurred: {e}")
        return None

# ウィンドウを最大化
def maximize_window():
    window = gw.getActiveWindow()
    if window:
        window.maximize()
    else:
        print("アクティブなウィンドウが見つかりませんでした。")

# アクティブウィンドウを保存して閉じる
def save_and_close_active_window():
    window = gw.getActiveWindow()
    if window:
        if 'Excel' in window.title:
            pyautogui.hotkey('ctrl', 's')
            time.sleep(1)
            window.close()
        else:
            window.close()
    else:
        print("アクティブなウィンドウが見つかりませんでした。")

# ファイルを削除
def delete_file(file_path):
    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"{file_path} を削除しました。")
    else:
        print(f"{file_path} が見つかりません。")

def delete_all_files_in_folder(folder_path):
    # フォルダが存在するか確認
    if os.path.exists(folder_path):
        # フォルダ内のすべてのファイルを削除
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)  # ファイルまたはリンクを削除
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)  # ディレクトリを削除
            except Exception as e:
                print(f'Failed to delete {file_path}. Reason: {e}')
    else:
        print(f'The folder {folder_path} does not exist.')

def rename_folder(old_folder_name, new_folder_name):
    try:
        os.rename(old_folder_name, new_folder_name)
        print(f"フォルダ名が '{old_folder_name}' から '{new_folder_name}' に変更されました。")
    except Exception as e:
        print(f"フォルダ名の変更中にエラーが発生しました: {e}")

def open_save_close_excel(file_path):
    # try:
    # Excelアプリケーションを起動
    excel_app = win32com.client.Dispatch("Excel.Application")
    excel_app.Visible = False  # Excelを表示しない
    excel_app.DisplayAlerts = False  # 警告を表示しない
    
    # ワークブックを開く
    workbook = excel_app.Workbooks.Open(file_path)
    
    # 変更を保存
    workbook.Save()
    workbook.Close(SaveChanges=True)
    excel_app.Quit()
        
    #     print("File opened, VBA executed, saved, and closed successfully.")
        
    # except Exception as e:
    #     print(f"An error occurred: {e}")

def modify_excel_k1(file_path, k1_value):
    try:
        # 一時ファイルのパスを作成
        temp_file_path = file_path.replace('.xlsx', '_temp.xlsx')
        
        # openpyxlを使用してExcelファイルを操作
        wb = load_workbook(file_path)
        sheet = wb.active
        sheet['K1'] = k1_value
        
        # 一時ファイルとして保存
        wb.save(temp_file_path)
        wb.close()
        
        # 一時ファイルを元のファイルに上書きコピー
        shutil.copy(temp_file_path, file_path)
        os.remove(temp_file_path)  # 一時ファイルを削除
        
        print("File opened, K1 value set, saved, and closed successfully.")
        
    except FileNotFoundError:
        print(file_path)
        print("File not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

def generate_combinations(start, end, group_size):
    # 数字の範囲を定義
    numbers = range(start, end + 1)

    # 重複組み合わせを生成
    combinations = [(i, j) for i in numbers for j in numbers if i <= j]

    # 組み合わせをグループに分ける
    groups = [combinations[i:i + group_size] for i in range(0, len(combinations), group_size)]

    # 最後のグループを補完するために、最初のグループから必要な数を追加
    if len(groups[-1]) < group_size:
        needed = group_size - len(groups[-1])
        groups[-1].extend(groups[0][:needed])

    # 各グループをDataFrameに変換
    df_list = []
    for idx, group in enumerate(groups):
        df = pd.DataFrame(group, columns=['Element1', 'Element2'])
        df_list.append(df)

    return df_list

def save_to_existing_excel(df, file_path, sheet_name, start_row, start_col):
    # 既存のエクセルファイルを読み込む
    book = load_workbook(file_path)
    sheet = book[sheet_name]

    # データを書き込む
    for i, row in df.iterrows():
        for j, value in enumerate(row):
            sheet.cell(row=start_row + i + 1, column=start_col + j + 1, value=value)

    # ファイルを保存
    book.save(file_path)

def copy_and_rename_condition_file(contingency_case, target_of_number, source_folder, destination_folder):
    if contingency_case == 1 and target_of_number == 1:
        src_file = "Condition(L)_1site.xlsx"
    elif contingency_case == 1 and target_of_number == 2:
        src_file = "Condition(L)_2site.xlsx"
    elif contingency_case == 2 and target_of_number == 1:
        src_file = "Condition(G_O_C)_1site.xlsx"
    elif contingency_case == 2 and target_of_number == 2:
        src_file = "Condition(G_O_C)_2site.xlsx"
    else:
        raise ValueError("Invalid combination of ContingencyCase and TargetOfNumber")
    
    src_path = os.path.join(source_folder, src_file)
    dst_path = os.path.join(destination_folder, "condition0.xlsx")
    
    if os.path.exists(src_path):
        shutil.copy(src_path, dst_path)
    else:
        raise FileNotFoundError(f"Source file {src_path} does not exist")

def copy_files2(main_folder, condition_list_folder, src_file_name, dst_file_name):
    src_mdl_path = os.path.join(condition_list_folder, src_file_name)
    dst_main_path = os.path.join(main_folder, dst_file_name)
    backup_mdl_path = os.path.join(condition_list_folder, dst_file_name)

    if os.path.exists(src_mdl_path):
        # 元のEAST10peak_AllThermal_AllDemand_auto.popをバックアップフォルダに移動
        if os.path.exists(dst_main_path):
            shutil.move(dst_main_path, backup_mdl_path)
        
        # MDLファイルをmainフォルダにコピー
        shutil.copy(src_mdl_path, dst_main_path)
        
        # ファイル名をEAST10peak_AllThermal_AllDemand_auto.popに変更
        os.rename(dst_main_path, os.path.join(main_folder, dst_file_name))

# メイン処理
if __name__ == "__main__":

    # 関数を使用して重複組み合わせを生成
    df_list = generate_combinations(11, 47, 10)
    # ファイルパスと書き込み設定
    sheet_name = 'Sheet1'  # 書き込みたいシート名
    start_row = 1
    start_col = 7

    ContingencyCase = 1 # 1: L, 2: GOC
    TargetOfNumber = 2  # 事故箇所数
    # NumberOfWork = 1   # 1ケースの作業回数（参考：TargetOfNumber=1->NumberOfWork15, TargetOfNumber=2->NumberOfWork100）

    if ContingencyCase == 1:
        k1_value = 1000
        k1_increment = 250
        k1_max = 1000
    elif ContingencyCase == 2:
        k1_value = 0.15
        k1_increment = 0.01
        k1_max = 0.15
    
    folder_path = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\02_MyModel\\CPATFreeR05V1.0\\CPATFree_R05_Ver1.0\\model\\EAST10\\main"
    file_name = "EAST10peak_AllThermal_AllDemand_auto.pop"
    path_to_folder = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\02_MyModel\\CPATFreeR05V1.0\\CPATFree_R05_Ver1.0\\Temp"
    condition_list_folder = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\02_MyModel\\CPATFreeR05V1.0\\CPATFree_R05_Ver1.0\\model\\EAST10\\List\\MDL"
    script_path = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\02_MyModel\\CPATFreeR05V1.0\\CPATFree_R05_Ver1.0\\model\\EAST10\\main\\MakeFile.py"
    excel_file_path = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\02_MyModel\\CPATFreeR05V1.0\\CPATFree_R05_Ver1.0\\model\\EAST10\\main\\condition0.xlsx"
    source_folder = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\02_MyModel\\CPATFreeR05V1.0\\CPATFree_R05_Ver1.0\\model\\EAST10\\List\\Condition"

    # ファイルのコピーと名前変更
    copy_and_rename_condition_file(ContingencyCase, TargetOfNumber, source_folder, folder_path)


    while k1_value <= k1_max:
        if ContingencyCase == 1:
            modify_excel_k1(excel_file_path, k1_value / TargetOfNumber)
        elif ContingencyCase == 2:
            modify_excel_k1(excel_file_path, k1_value)
        

        mdl_files = ["M2.pop"]
        # mdl_files = ["M8.pop", "M2.pop", "M16.pop"]
        for mdl_file in mdl_files:
            copy_files2(folder_path, condition_list_folder, mdl_file, file_name)
            mdl_base_name = os.path.splitext(mdl_file)[0]
            
            for i, df in enumerate(df_list):

                # 既存のエクセルファイルにデータを書き込む
                save_to_existing_excel(df, excel_file_path, sheet_name, start_row, start_col)
                
                copy_file(folder_path, file_name)

                action_files = ["OpenCPAT.json"]
                replay_all_actions(action_files)

                maximize_window()

                open_save_close_excel(excel_file_path)
                execute_python_script(script_path)
                b3_value = i

                # ContingencyCaseに応じてclipboard_stringを変更
                if ContingencyCase == 1:
                    clipboard_string = f"00{mdl_base_name}-{int(k1_value)}MW-{b3_value}"
                elif ContingencyCase == 2:
                    clipboard_string = f"00{mdl_base_name}-{int(k1_value*1000)}msec-{b3_value}"
                else:
                    raise ValueError("ContingencyCaseは1または2でなければなりません。")
                
                pyperclip.copy(clipboard_string)
                
                # b3_value = operate_excel(excel_file_path, mdl_base_name, k1_value, ContingencyCase)

                if b3_value is None:
                    continue

                action_files = ["ExecuteCPAT.json"]
                replay_all_actions(action_files)

                save_and_close_active_window()
                time.sleep(1)

                action_files = ["CloseCPAT.json"]
                replay_all_actions(action_files)

                delete_file(os.path.join(folder_path, "EAST10peak_AllThermal_AllDemand_auto_copy.pop"))
                # delete_all_files_in_folder(os.path.join(path_to_folder, f"EAST10peak_AllThermal_AllDemand_auto_copy_{b3_value}"))
                rename_folder(os.path.join(path_to_folder, "EAST10peak_AllThermal_AllDemand_auto_copy"), os.path.join(path_to_folder, str(clipboard_string)))
                
        k1_value += k1_increment
