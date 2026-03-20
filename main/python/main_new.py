import shutil
import os
import time
import json
from pynput import mouse, keyboard
import pygetwindow as gw
import pyautogui
import subprocess
import pyperclip
import math
from openpyxl import load_workbook
import pandas as pd
import win32com.client
from itertools import combinations_with_replacement
import zipfile
import numpy as np
import subprocess

# ファイルの名前を変更する関数
def rename_files(original_file, new_file):
    os.rename(original_file, new_file)

# ディレクトリを作成する関数
def create_directory(directory):
    os.makedirs(directory, exist_ok=True)

# ZIPファイルを解凍する関数
def unzip_file(zip_file, extract_to):
    import subprocess
    import os

    try:
        # PowerShellのExpand-Archiveコマンドを使用してZIPを解凍
        powershell_command = f'Expand-Archive -Path "{zip_file}" -DestinationPath "{extract_to}" -Force'
        subprocess.run(['powershell', '-Command', powershell_command], check=True)
    except subprocess.CalledProcessError as e:
        print(f"Error extracting {zip_file}: {str(e)}")
        raise
    except Exception as e:
        print(f"Unexpected error: {str(e)}")
        raise

# ディレクトリ内の特定のファイルを削除する関数
def delete_files(directory, extension):
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith(extension):
                os.remove(os.path.join(root, file))

# ファイルを移動する関数
def move_files(source_directory, destination_directory, extension):
    for root, dirs, files in os.walk(source_directory):
        for file in files:
            if file.endswith(extension):
                shutil.move(os.path.join(root, file), destination_directory)

# ファイルの内容を置換する関数
def replace_content_in_files(directory, target, replacement, extension):
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            if file.endswith(extension):
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                modified_content = content.replace(target, replacement)
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(modified_content)

# valueを計算する関数
def calculate_value(factor, res):
    return factor * res

# Demand.pnsj ファイルの内容を置換する関数
def modify_demand_file(file_path, value_text_pgo, value_text_qgo):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    modified_content = content.replace("<Pgo>0.0</Pgo>", value_text_pgo)
    modified_content = modified_content.replace("<Qgo>0.0</Qgo>", value_text_qgo)
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(modified_content)

def modify_demand_file_GOC(file_path, sc_posi_value, new_use_name="運用"):
    # ファイルの内容を読み込む
    with open(file_path, "r", encoding="utf-8") as file:
        lines = file.readlines()
    # CodeNumber の位置を探す
    target_index = None
    for i, line in enumerate(lines):
        if f"<CodeNumber>{sc_posi_value}</CodeNumber>" in line:
            target_index = i
            break
    # 2行上の UseName を修正
    if target_index is not None and target_index >= 2:
        for j in range(target_index - 2, target_index):
            if "<UseName>" in lines[j]:
                lines[j] = f"    <UseName>{new_use_name}</UseName>\n"
                break
    # 修正後のファイルを上書き保存
    with open(file_path, "w", encoding="utf-8") as file:
        file.writelines(lines)

# <Pgo>タグと<Qgo>タグを新しい値で置き換える関数
def replace_tags_in_file(file_path, old_values_pgo, new_values_pgo, old_values_qgo_temp, new_values_qgo):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    for old_value, new_value in zip(old_values_pgo, new_values_pgo):
        old_value_text = f"<Pgo>{old_value}</Pgo>"
        new_value_text = f"<Pgo>{new_value}</Pgo>"
        content = content.replace(old_value_text, new_value_text)
    for old_value, new_value in zip(old_values_qgo_temp, new_values_qgo):
        old_value_text = f"<Qgo>{old_value}</Qgo>"
        new_value_text = f"<Qgo>{new_value}</Qgo>"
        content = content.replace(old_value_text, new_value_text)
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(content)

# Yco 値を変更する関数
def modify_sc_mva_value(file_path, sc_posi_value, sc_mva_value):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    lines = content.split('\\n')
    for i, line in enumerate(lines):
        if f"<Number>{sc_posi_value}</Number>" in line:
            if i + 7 < len(lines) and "<Yco>" in lines[i + 7]:
                lines[i + 7] = f"          <Yco>{sc_mva_value}</Yco>"
    with open(file_path, 'w', encoding='utf-8') as file:
        file.write('\\n'.join(lines))
    print(f"{os.path.basename(file_path)} の <Number>{sc_posi_value}</Number> の部分の <Yco> を {sc_mva_value} に変更しました。")

# ディレクトリを圧縮する関数
def compress_directory(directory, output_file):
    shutil.make_archive(output_file.replace('.zip', ''), 'zip', directory)

# 中間ファイルと作業ディレクトリを削除する関数
def clean_up(*files_and_directories):
    for item in files_and_directories:
        if os.path.isfile(item):
            os.remove(item)
        elif os.path.isdir(item):
            shutil.rmtree(item)

# ファイルをコピー
def copy_file(folder_path, file_name):
    source_path = os.path.join(folder_path, file_name)
    destination_path = os.path.join(folder_path, file_name.replace(".pop", "_copy.pop"))
    shutil.copy2(source_path, destination_path)
    print(f"ファイル {file_name} を {destination_path} にコピーしました。")

def execute_python_script(script_path):
    # try:
    # サブプロセスを使ってスクリプトを実行
    result = subprocess.run(['python', script_path], capture_output=True, text=True)
    
    # 実行結果の表示
    if result.returncode == 0:
        print("Script executed successfully.")
        print("Output:\\n", result.stdout)
    else:
        print("Script execution failed.")
        print("Error:\\n", result.stderr)
    
    # except Exception as e:
    #     print(f"An error occurred: {e}")

# マウスとキーボードのアクションを再生
def replay_actions(file_name):
    with open(file_name, 'r') as f:
        actions = json.load(f)

    start_time = actions[0][-1]
    mouse_controller = mouse.Controller()
    keyboard_controller = keyboard.Controller()

    for action in actions:
        action_type = action[0]
        if action_type == 'move':
            _, x, y, timestamp = action
            mouse_controller.position = (x, y)
        elif action_type == 'click':
            _, x, y, button, pressed, timestamp = action
            mouse_controller.position = (x, y)
            button = mouse.Button[button.split('.')[1]]
            if pressed:
                mouse_controller.press(button)
            else:
                mouse_controller.release(button)
        elif action_type == 'scroll':
            _, x, y, dx, dy, timestamp = action
            mouse_controller.scroll(dx, dy)
        elif action_type == 'key':
            _, key, event_type, timestamp = action
            try:
                if key.startswith('Key.'):
                    key = getattr(keyboard.Key, key.split('.')[1])
                if event_type == 'press':
                    keyboard_controller.press(key)
                else:
                    keyboard_controller.release(key)
            except AttributeError:
                print(f"未対応のキー: {key}")
            except ValueError as e:
                print(f"キーの解決中にエラーが発生しました: {e}")

        time.sleep(max(0, timestamp - start_time))
        start_time = timestamp

def replay_all_actions(file_names):
    for file_name in file_names:
        print(f"再生中: {file_name}")
        replay_actions(file_name)
        print(f"再生完了: {file_name}")

# Excelファイルを操作
def operate_excel(file_path, mdl_base_name, fault_value, ContingencyCase):
    try:
        # openpyxlを使ってExcelファイルを読み込む
        from openpyxl import load_workbook
        import math
        import pyperclip

        workbook = load_workbook(file_path)
        sheet = workbook.active
        b3_value = math.floor(sheet["O2"].value)
        print(f"セルO2の値: {b3_value}")
        workbook.save(file_path)
        
        # ContingencyCaseに応じてclipboard_stringを変更
        if ContingencyCase == 1:
            clipboard_string = f"00{mdl_base_name}-{int(fault_value)}MW-{b3_value}"
        elif ContingencyCase == 2:
            clipboard_string = f"00{mdl_base_name}-{int(fault_value*1000)}msec-{b3_value}"
        else:
            raise ValueError("ContingencyCaseは1または2でなければなりません。")
        
        pyperclip.copy(clipboard_string)
        print(f"文字列 '{clipboard_string}' をクリップボードにコピーしました。")
        return b3_value
    except Exception as e:
        print(f"An error occurred: {e}")
        return None

# ウィンドウを最大化
def maximize_window():
    window = gw.getActiveWindow()
    if window:
        window.maximize()
    else:
        print("アクティブなウィンドウが見つかりませんでした。")

# アクティブウィンドウを保存して閉じる
def save_and_close_active_window():
    window = gw.getActiveWindow()
    if window:
        if 'Excel' in window.title:
            pyautogui.hotkey('ctrl', 's')
            time.sleep(1)
            window.close()
        else:
            window.close()
    else:
        print("アクティブなウィンドウが見つかりませんでした。")

# ファイルを削除
def delete_file(file_path):
    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"{file_path} を削除しました。")
    else:
        print(f"{file_path} が見つかりません。")

def delete_all_files_in_folder(folder_path):
    # フォルダが存在するか確認
    if os.path.exists(folder_path):
        # フォルダ内のすべてのファイルを削除
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)  # ファイルまたはリンクを削除
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)  # ディレクトリを削除
            except Exception as e:
                print(f'Failed to delete {file_path}. Reason: {e}')
    else:
        print(f'The folder {folder_path} does not exist.')

def rename_folder(old_folder_name, new_folder_name):
    try:
        os.rename(old_folder_name, new_folder_name)
        print(f"フォルダ名が '{old_folder_name}' から '{new_folder_name}' に変更されました。")
    except Exception as e:
        print(f"フォルダ名の変更中にエラーが発生しました: {e}")

def open_save_close_excel(file_path):
    # try:
    # Excelアプリケーションを起動
    excel_app = win32com.client.Dispatch("Excel.Application")
    excel_app.Visible = False  # Excelを表示しない
    excel_app.DisplayAlerts = False  # 警告を表示しない
    
    # ワークブックを開く
    workbook = excel_app.Workbooks.Open(file_path)
    
    # 変更を保存
    workbook.Save()
    workbook.Close(SaveChanges=True)
    excel_app.Quit()
        
    #     print("File opened, VBA executed, saved, and closed successfully.")
        
    # except Exception as e:
    #     print(f"An error occurred: {e}")

def modify_excel_k1(file_path, fault_value, PL_value, sc_mva_value, sc_posi_value):
    try:
        # 一時ファイルのパスを作成
        temp_file_path = file_path.replace('.xlsx', '_temp.xlsx')
        
        # openpyxlを使用してExcelファイルを操作
        wb = load_workbook(file_path)
        sheet = wb.active
        sheet['O1'] = fault_value
        sheet['T4'] = PL_value
        sheet['T5'] = sc_mva_value
        sheet['T6'] = sc_posi_value
        
        # 一時ファイルとして保存
        wb.save(temp_file_path)
        wb.close()
        
        # 一時ファイルを元のファイルに上書きコピー
        shutil.copy(temp_file_path, file_path)
        os.remove(temp_file_path)  # 一時ファイルを削除
        
        print("File opened, K1 value set, saved, and closed successfully.")
    except FileNotFoundError:
        print(file_path)
        print("File not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

def modify_excel_GOC(file_path, fault_value, PL_value, sc_mva_value, sc_posi_value):
    try:
        # 一時ファイルのパスを作成
        temp_file_path = file_path.replace('.xlsx', '_temp.xlsx')
        
        # openpyxlを使用してExcelファイルを操作
        wb = load_workbook(file_path)
        sheet = wb.active
        sheet['O1'] = fault_value
        sheet['T4'] = PL_value
        sheet['T5'] = sc_mva_value
        sheet['T6'] = sc_posi_value
        
        # 一時ファイルとして保存
        wb.save(temp_file_path)
        wb.close()
        
        # 一時ファイルを元のファイルに上書きコピー
        shutil.copy(temp_file_path, file_path)
        os.remove(temp_file_path)  # 一時ファイルを削除
        
        print("File opened, K1 value set, saved, and closed successfully.")
        
    except FileNotFoundError:
        print(file_path)
        print("File not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

def generate_combinations(N_LIST, group_size, i):
    # 係数分類時の学習データ作成
    ## 全ケース（逆順も含む）
    if i == 2:
        # i=2の場合、同じ数字の組み合わせを除外
        combinations = [c for c in combinations_with_replacement(N_LIST, i) if c[0] != c[1]]
        combinations.extend([tuple(reversed(c)) for c in combinations if c[0] != c[1]])
    ## ランダムに300パターンを抽出（ただし、組み合わせの総数が300未満の場合は全パターンを使用）
    import random
    if len(combinations) > 300:
        combinations = random.sample(combinations, 300)
    
    # 全ケースの検証
    # combinations = list(combinations_with_replacement(N_LIST, i))

    # 組み合わせをグループに分ける
    groups = [combinations[j:j + group_size] for j in range(0, len(combinations), group_size)]

    # 最後のグループを補完するために、最初のグループから必要な数を追加
    if len(groups[-1]) < group_size:
        needed = group_size - len(groups[-1])
        groups[-1].extend(groups[0][:needed])

    # 各グループをDataFrameに変換
    df_list = []
    for idx, group in enumerate(groups):
        df = pd.DataFrame(group, columns=[f'Element{k+1}' for k in range(i)])
        df_list.append(df)

    return df_list

def save_to_existing_excel(df, file_path, sheet_name, start_row, start_col):
    # 既存のエクセルファイルを読み込む
    book = load_workbook(file_path)
    sheet = book[sheet_name]

    # データを書き込む
    for i, row in df.iterrows():
        for j, value in enumerate(row):
            sheet.cell(row=start_row + i + 1, column=start_col + j + 1, value=value)

    # ファイルを保存
    book.save(file_path)

def copy_and_rename_condition_file(contingency_case, target_of_number, source_folder, destination_folder):
    if contingency_case == 1 and target_of_number == 1:
        src_file = "Condition(L)_1site.xlsx"
    elif contingency_case == 1 and target_of_number == 2:
        src_file = "Condition(L)_2site.xlsx"
    elif contingency_case == 2 and target_of_number == 1:
        src_file = "Condition(G_O_C)_1site.xlsx"
    elif contingency_case == 2 and target_of_number == 2:
        src_file = "Condition(G_O_C)_2site.xlsx"
    else:
        raise ValueError("Invalid combination of ContingencyCase and TargetOfNumber")
    
    src_path = os.path.join(source_folder, src_file)
    dst_path = os.path.join(destination_folder, "condition0.xlsx")
    
    if os.path.exists(src_path):
        shutil.copy(src_path, dst_path)
    else:
        raise FileNotFoundError(f"Source file {src_path} does not exist")

def copy_files2(main_folder, condition_list_folder, src_file_name, dst_file_name):
    src_mdl_path = os.path.join(condition_list_folder, src_file_name)
    dst_main_path = os.path.join(main_folder, dst_file_name)
    backup_mdl_path = os.path.join(condition_list_folder, dst_file_name)

    if os.path.exists(src_mdl_path):
        # 元のEAST10peak_AllThermal_AllDemand_auto.popをバックアップフォルダに移動
        if os.path.exists(dst_main_path):
            shutil.move(dst_main_path, backup_mdl_path)
        
        # MDLファイルをmainフォルダにコピー
        shutil.copy(src_mdl_path, dst_main_path)
        
        # ファイル名をEAST10peak_AllThermal_AllDemand_auto.popに変更
        os.rename(dst_main_path, os.path.join(main_folder, dst_file_name))

def is_process_running(process_name):
    try:
        # 実行中のプロセスリストを取得
        result = subprocess.run(['tasklist', '/FI', f'IMAGENAME eq {process_name}'], capture_output=True, text=True)
        # プロセス名が結果に含まれているか確認
        return process_name in result.stdout
    except Exception as e:
        print(f"Error checking process: {e}")
        return False

def get_process_ids(process_name):
    result = subprocess.run(['tasklist'], capture_output=True, text=True)
    lines = result.stdout.splitlines()
    pids = []

    for line in lines:
        if process_name.lower() in line.lower():

            
            parts = line.split()
            pid = int(parts[1])
            pids.append(pid)
    
    return pids

def kill_process_v1(pid):
    subprocess.run(['taskkill', '/PID', str(pid), '/F'])

def terminate_process(process_name):
    try:
        # WMIサービスに接続
        wmi = win32com.client.GetObject('winmgmts:')
        
        # 指定されたプロセス名を持つすべてのプロセスを列挙
        processes = wmi.ExecQuery(f"SELECT * FROM Win32_Process WHERE Name='{process_name}'")
        
        for process in processes:
            print(f"Terminating process: {process.Name} (PID: {process.ProcessId})")
            process.Terminate()
            
        if not processes:
            print(f"No processes found with name: {process_name}")
    except Exception as e:
        print(f"Failed to terminate process: {e}")

def restart_explorer():
    try:
        subprocess.run('start explorer.exe', shell=True, check=True)
        print('Explorer has been restarted.')
    except subprocess.CalledProcessError as e:
        print(f'Failed to restart Explorer. Error: {e}')

def generate_combinations_all(RES_LIST, SG_M_LIST, SC_MVA_LIST, SC_POSI_LIST, FAULT_POSI_LIST, FAULT_LIST):
    combinations = []
    for fault_value in FAULT_LIST: # 事故量
        for res_value in RES_LIST: # 再エネ接続地点
            for m_value in SG_M_LIST: # 同期発電機の慣性定数
                for sc_mva_value in SC_MVA_LIST: # 調相設備の容量
                    for sc_posi_value in SC_POSI_LIST: # 調相設備の接続位置
                        for i, _ in enumerate(FAULT_POSI_LIST):
                            combinations.append((res_value, m_value, sc_mva_value, sc_posi_value, i, fault_value))
                            # 再エネ、慣性、調相設備容量、調相設備位置、番目（事故位置のリスト番号）、事故量
    return combinations

def create_combinations(TargetOfNumber, k1_start, k1_end, delta_k1, FAULT_POSI_LIST):
    RES_LIST = [0.8, 0.5]
    SG_M_LIST = [2, 8, 16]
    SC_MVA_LIST = [0.25]
    SC_POSI_LIST = range(11, 48)

    if k1_start == 250:
        FAULT_LIST = list(range(k1_start, k1_end, delta_k1))
    elif k1_start == 1:
        fault_values = [x / 100 for x in list(range(k1_start, k1_end, delta_k1))]

    combinations = generate_combinations_all(RES_LIST, SG_M_LIST, SC_MVA_LIST, SC_POSI_LIST, FAULT_POSI_LIST, FAULT_LIST)
    # combinations = generate_combinations_all(res_list, M_list, yoc_list, bus_list, df_list, fault_values)
    return combinations

def main(start_combination, combinations, sheet_name, start_row, start_col, PL_base, ContingencyCase, TargetOfNumber, folder_path, path_to_folder, condition_list_folder, script_path, excel_file_path, source_folder, source, destination_directory, original_file_sys, new_file_sys, new_directory_sys, new_zip_file_sys, original_file_condition, new_file_condition, new_directory_condition, new_zip_file_condition, Lp_i, Lq_i, df_list):
    # ファイル名
    file_name = os.path.basename(source)

    # コピー先のファイルパス
    destination = os.path.join(destination_directory, file_name)

    # Excelファイルのコピーと名前変更
    copy_and_rename_condition_file(ContingencyCase, TargetOfNumber, source_folder, folder_path)

    # 指定された組み合わせの次の行から開始
    try:
        start_index = combinations.index(start_combination)
    except ValueError:
        start_index = 0  # 組み合わせが見つからなかった場合は最初から

    remaining_combinations = combinations[start_index:]
    print(len(remaining_combinations)*2.5/60) # 1ケース2.5分で計算 [時間]
    # remaining_combinations = combinations[start_index:]

    # for i in range(1):

    for combination in remaining_combinations:
        action_files = ["OPEN_CPAT.json"]
        replay_all_actions(action_files)
        
        # combination = remaining_combinations
        res_value, m_value, sc_mva_value, sc_posi_value, i, fault_value = combination
        replacement_text = f"<Mg>{m_value}</Mg>"

        value_pgo = calculate_value(Lp_i, res_value)
        value_qgo = calculate_value(Lq_i, res_value)
        value_text_pgo = f"<Pgo>{value_pgo}</Pgo>"
        value_text_qgo = f"<Qgo>{value_qgo}</Qgo>"

        Pgo = np.array([5.97560976, 9.3902439, 5.12195122, 9.3902439, 5.12195122, 9.3902439, 9.3902439, 5.97560976, 5.97560976, 4.26829268])
        Qgo_temp = np.array([2.47261939, 3.88554476, 2.11938805, 3.88554476, 2.11938805, 3.88554476, 3.88554476, 2.47261939, 2.47261939, 1.76615671])
        Qgo = np.array([2.66181, 3.86776, 1.45165, 3.7865, 1.6436, 3.97747, 5.75334, 1.73422, 1.82973, 2.25889])

        Pgo_total = np.sum(Pgo)
        Qgo_total = np.sum(Qgo)

        Pgo_percentages = (Pgo / Pgo_total)

        Pgo_res = calculate_value(Lp_i, res_value) * 37
        Qgo_res = calculate_value(Lq_i, res_value) * 37

        Pgo_res = Pgo - (Pgo_res * Pgo_percentages)
        Qgo_res = (Qgo_total - Qgo_res) * Pgo_percentages

        PL_value = PL_base * (1 - (res_value / 100))

        # Excelファイルの値更新
        if ContingencyCase == 1:
            modify_excel_k1(excel_file_path, fault_value / TargetOfNumber, PL_value, sc_mva_value, sc_posi_value)
        # elif ContingencyCase == 2: GOCは地絡であるため、事故量は関係ないから、個々での操作は不要
        #     modify_excel_GOC(excel_file_path, fault_value, PL_value)

        df = df_list[i]
        # 既存のエクセルファイルにデータを書き込む
        save_to_existing_excel(df, excel_file_path, sheet_name, start_row, start_col)
        
        # 条件設定
        open_save_close_excel(excel_file_path)
        execute_python_script(script_path)
        b3_value = i

        # 系統コピー
        shutil.copy2(source, destination)

        rename_files(original_file_sys, new_file_sys)
        rename_files(original_file_condition, new_file_condition)

        # 新しいディレクトリを作成
        create_directory(new_directory_sys)
        create_directory(new_directory_condition)

        # ZIPファイルを解凍
        unzip_file(new_file_sys, new_directory_sys)
        unzip_file(new_file_condition, new_directory_condition)

        # ディレクトリ内の .pnsy ファイルを削除
        delete_files(new_directory_sys, ".pnsy")

        # condition0_unzipped 内の .pnsy ファイルを EAST10peak_AllThermal_AllDemand_auto_res_SPM_unzipped に移動
        move_files(new_directory_condition, new_directory_sys, ".pnsy")

        # 解凍されたファイルを処理（？）
        replace_content_in_files(new_directory_sys, "<Mg>2.0</Mg>", replacement_text, ".pnsd")

        # Demand.pnsj ファイルを変更
        if ContingencyCase == 1:
            file_path = os.path.join(new_directory_sys, "Demand.pnsj")
            modify_demand_file(file_path, value_text_pgo, value_text_qgo)
            # <Pgo>タグと<Qgo>タグを新しい値で置き換え
            replace_tags_in_file(file_path, Pgo, Pgo_res, Qgo_temp, Qgo_res)
            # Yco 値を変更
            modify_sc_mva_value(file_path, sc_posi_value, sc_mva_value)
        elif ContingencyCase == 2:
            # 新しい拡張子を付ける
            new_file_name = "275-500kVloop_EAST10peak_10SG_3RES_1SC" + ".pnsw"
            file_path = os.path.join(new_directory_sys, new_file_name)
            modify_demand_file_GOC(file_path, sc_posi_value)

        # ディレクトリを圧縮
        compress_directory(new_directory_sys, new_zip_file_sys)

        # 圧縮したファイルの名前を元に戻す
        rename_files(new_zip_file_sys, original_file_sys)

        # 作業ディレクトリと中間ファイルを削除
        clean_up(new_file_sys, new_file_condition, new_directory_sys, new_directory_condition)

        # ContingencyCaseに応じてclipboard_stringを変更
        if ContingencyCase == 1:
            clipboard_string = f"M{m_value}-RES{res_value*100}%-YCbus{int(sc_posi_value)}-{int(sc_mva_value * 1000)}MVA-ΔP{int(fault_value)}MW-{b3_value}"
        elif ContingencyCase == 2:
            clipboard_string = f"SynC_POSI_{sc_posi_value}-FAULT_num_{b3_value}"
        else:
            raise ValueError("ContingencyCaseは1または2でなければなりません。")
        pyperclip.copy(clipboard_string)

        action_files = ["OPEN_MDL.json"]
        replay_all_actions(action_files)
    
        action_files = ["SET_CONDITION.json"]
        replay_all_actions(action_files)
        # time.sleep(10)

        action_files = ["RESET_CONDITION.json"]
        replay_all_actions(action_files)

        action_files = ["EXECUTE_ANAYSIS.json"]
        replay_all_actions(action_files)

        action_files = ["SAVE_CSV.json"]
        replay_all_actions(action_files)
        
        # CPATとExcelのプロセスIDを取得し，終了
        cpat_pids = get_process_ids('CPAT.exe')
        excel_pids = get_process_ids('EXCEL.EXE')
        for pid in cpat_pids:
            kill_process_v1(pid)
        
        for pid in excel_pids:
            kill_process_v1(pid)
        
        # explorer_name = 'explorer.exe'
        # if is_process_running(explorer_name):
        #     kill_process_v2(explorer_name)
        #     # explorer.exeを新しいタスクとして実行
        #     start_process(explorer_name)
        # else:
        #     print(f"{explorer_name} process not found.")
        # terminate_process('explorer.exe')
        # restart_explorer()

        delete_file(os.path.join(folder_path, "275-500kVloop_EAST10peak_10SG_3RES_1SC_copy.pop"))
        rename_folder(os.path.join(path_to_folder, "275-500kVloop_EAST10peak_10SG_3RES_1SC"), os.path.join(path_to_folder, str(clipboard_string)))
        os.remove(destination)

if __name__ == "__main__":
    # 諸々設定
    sheet_name = 'Sheet1'  # 書き込みたいシート名
    start_row = 1
    start_col = 11
    PL_base = 1.89189189
    ContingencyCase = 2  # 1: 負荷増加, 2: 短絡⇒遮断⇒閉路
    TargetOfNumber = 2  # 事故箇所数

    if ContingencyCase == 1:
        RES_LIST = [0.8, 0.5]
        SG_M_LIST = [2, 8, 16]
        SC_POSI_LIST = range(11, 48)
        SC_MVA_LIST = [0.25]
        N_LIST = range(11, 47 + 1) # 対象ノードは11~47の間
        FAULT_POSI_LIST = generate_combinations(N_LIST, 10, TargetOfNumber)
        FAULT_LIST = list(range(250, 1001, 250))
    elif ContingencyCase == 2:
        RES_LIST = [0]
        SG_M_LIST = [8]
        SC_MVA_LIST = [0]
        SC_POSI_LIST = pd.read_csv("C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\03_Paid\\R06_1.0_GUI_Pub_福井大学（電力システム研究室）\\CPATPubR06V1.0\\CPATPub\\main\\python\\G_O_C_SC_POSI_LIST.csv")
        SC_POSI_LIST = SC_POSI_LIST.iloc[:, 0].dropna().astype(int).tolist() # 調相設備が接続されているブランチのリスト
        SC_POSI_LIST = [0]  # ダミーの値を設定
        N_LIST = pd.read_csv("C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\03_Paid\\R06_1.0_GUI_Pub_福井大学（電力システム研究室）\\CPATPubR06V1.0\\CPATPub\\main\\python\\G_O_C_FAULT_POSI_LIST.csv")
        N_LIST = N_LIST.iloc[:, 0].dropna().astype(int).tolist()
        FAULT_POSI_LIST = generate_combinations(N_LIST, 10, TargetOfNumber)
        FAULT_LIST = [1]

    folder_path = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\03_Paid\\R06_1.0_GUI_Pub_福井大学（電力システム研究室）\\CPATPubR06V1.0\\CPATPub\\main\\python"
    path_to_folder = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\03_Paid\\R06_1.0_GUI_Pub_福井大学（電力システム研究室）\\CPATPubR06V1.0\\CPATPub\\Temp"
    condition_list_folder = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\03_Paid\\R06_1.0_GUI_Pub_福井大学（電力システム研究室）\\CPATPubR06V1.0\\CPATPub\\main\\mdl"
    script_path = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\03_Paid\\R06_1.0_GUI_Pub_福井大学（電力システム研究室）\\CPATPubR06V1.0\\CPATPub\\main\\python\\MakeFile.py"
    excel_file_path = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\03_Paid\\R06_1.0_GUI_Pub_福井大学（電力システム研究室）\\CPATPubR06V1.0\\CPATPub\\main\\python\\condition0.xlsx"
    source_folder = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\03_Paid\\R06_1.0_GUI_Pub_福井大学（電力システム研究室）\\CPATPubR06V1.0\\CPATPub\\main\\condition"
    # 起動停止状態：sourceのpopファイルを、手動で変更する
    source = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\03_Paid\\R06_1.0_GUI_Pub_福井大学（電力システム研究室）\\CPATPubR06V1.0\\CPATPub\\main\\mdl\\275-500kVloop_EAST10peak_10SG_3RES_1SC.pop"
    destination_directory = "C:\\Users\\PowerSystemLab\\Documents\\FromDesktop\\01_研究資料\\05_実行ファイル\\CPAT\\03_Paid\\R06_1.0_GUI_Pub_福井大学（電力システム研究室）\\CPATPubR06V1.0\\CPATPub\\main\\python"

    # 定数定義
    original_file_sys = "275-500kVloop_EAST10peak_10SG_3RES_1SC.pop"
    new_file_sys = "new_file_sys.zip"
    new_directory_sys = "new_file_sys_unzipped"
    new_zip_file_sys = "new_file_sys_new.zip"

    original_file_condition = "condition0.pop"
    new_file_condition = "condition0.zip"
    new_directory_condition = "condition0_unzipped"
    new_zip_file_condition = "condition0_new.zip"

    Lp_i = 1.89189189
    Lq_i = 0.66824879

    # # ファイルを指定する場合
    # combinations = (
    #     (0.8, 16, 0.25, 34, 0, 750),
    # )

    combinations = generate_combinations_all(RES_LIST, SG_M_LIST, SC_MVA_LIST, SC_POSI_LIST, FAULT_POSI_LIST, FAULT_LIST)
    combinations = sorted(combinations, key=lambda x: x[4])
    # combinations = combinations[0:3]

    # combinations = create_combinations(TargetOfNumber, k1_start, k1_end, delta_k1, df_list)
start_combination = combinations[0]
    # for start_combination in combinations:
        # start_combination = (0, 8, 0, 0, 51, 1) # 1: (0.8, 16, 0.25, 47, 0, 500) (res, m_value, sc_mva_value, sc_posi_value, df, fault_value)
        # # M16-RES80.0%-YCbus27-250MVA-ΔP500MW-0
main(start_combination, combinations, sheet_name, start_row, start_col, PL_base, ContingencyCase, TargetOfNumber, folder_path, path_to_folder, condition_list_folder, script_path, excel_file_path, source_folder, source, destination_directory, original_file_sys, new_file_sys, new_directory_sys, new_zip_file_sys, original_file_condition, new_file_condition, new_directory_condition, new_zip_file_condition, Lp_i, Lq_i, FAULT_POSI_LIST)